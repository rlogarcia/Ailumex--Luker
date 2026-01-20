# -*- coding: utf-8 -*-
"""
Wizard de Importación Masiva de Estudiantes + Matrículas
=========================================================

Especificación Técnica:
- Importa desde Excel con columnas específicas
- Crea/actualiza estudiantes según CÓDIGO USUARIO
- Genera matrículas coherentes con el modelo académico
- Asigna asistencia histórica según campo NIVEL
- Procesa congelamientos según DÍAS CONG.
- Aplica estados académicos finales

Categorías Permitidas: B teens, ADULTOS
Programa: B teens → B teens, ADULTOS → Benglish
Plan: GOLD → PLAN GOLD, PLUS → PLAN PLUS, etc.
"""

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
from datetime import datetime, timedelta
import base64
import io
import re
import logging

try:
    import openpyxl
except ImportError:
    openpyxl = None

_logger = logging.getLogger(__name__)
# Cache de tipos de identificación a nivel de módulo: evita añadir atributos
# a instancias de recordset (que pueden fallar en algunos entornos/serializaciones).
_ID_TYPE_CACHE = {}


class StudentEnrollmentImportWizard(models.TransientModel):
    """Wizard para importación masiva de estudiantes y matrículas desde Excel"""

    _name = "benglish.student.enrollment.import.wizard"
    _description = "Importación Masiva de Estudiantes y Matrículas"

    # ========================================================================
    # CAMPOS
    # ========================================================================

    file_data = fields.Binary(
        string="Archivo Excel",
        required=True,
        help="Archivo Excel (.xlsx) con los datos de estudiantes y matrículas",
    )
    file_name = fields.Char(string="Nombre del Archivo", required=True)

    # Configuración
    update_existing = fields.Boolean(
        string="Actualizar Existentes",
        default=True,
        help="Si está marcado, actualiza estudiantes existentes. Si no, solo crea nuevos.",
    )
    skip_errors = fields.Boolean(
        string="Omitir Errores",
        default=False,
        help="Si está marcado, continúa la importación aunque haya errores en algunas filas",
    )

    # Resultados
    state = fields.Selection(
        [
            ("draft", "Borrador"),
            ("processing", "Procesando"),
            ("done", "Completado"),
            ("error", "Error"),
        ],
        default="draft",
        string="Estado",
    )

    log_ids = fields.One2many(
        "benglish.student.enrollment.import.log",
        "wizard_id",
        string="Log de Importación",
    )

    total_rows = fields.Integer(string="Total Filas", readonly=True)
    success_count = fields.Integer(string="Importados Exitosamente", readonly=True)
    error_count = fields.Integer(string="Errores", readonly=True)
    skipped_count = fields.Integer(string="Omitidos", readonly=True)

    # ========================================================================
    # COLUMNAS EXACTAS DEL EXCEL (según especificación)
    # ========================================================================

    EXPECTED_COLUMNS = [
        "CÓDIGO USUARIO",
        "PRIMER NOMBRE",
        "SEGUNDO NOMBRE",
        "PRIMER APELLIDO",
        "SEGUNDO APELLIDO",
        "EMAIL",
        "DOCUMENTO",
        "CATEGORÍA",
        "PLAN",
        "SEDE",
        "F. INICIO CURSO",
        "DÍAS CONG.",
        "FECHA FIN CURSO MÁS CONG.",
        "FASE",
        "NIVEL",
        "ESTADO",
        "CONTACTO TÍTULAR",
        "FECHA NAC.",
    ]

    HEADER_ALIASES = {
        "CODIGOUSUARIO": "CÓDIGO USUARIO",
        "CODIGOEST": "CÓDIGO USUARIO",
        "CODIGOESTUDIANTE": "CÓDIGO USUARIO",
        "PRIMERNOMBRE": "PRIMER NOMBRE",
        "SEGUNDONOMBRE": "SEGUNDO NOMBRE",
        "PRIMERAPELLIDO": "PRIMER APELLIDO",
        "PIRMERAPELLIDO": "PRIMER APELLIDO",
        "SEGUNDOAPELLIDO": "SEGUNDO APELLIDO",
        "EMAIL": "EMAIL",
        "EMAILEST": "EMAIL",
        "CORREO": "EMAIL",
        "DOCUMENTO": "DOCUMENTO",
        "IDENTIFICACION": "DOCUMENTO",
        "NUMERODOCUMENTO": "DOCUMENTO",
        "TIPOIDENTIFICACION": "TIPO IDENTIFICACION",
        "TIPODOCUMENTO": "TIPO IDENTIFICACION",
        "TIPODOC": "TIPO IDENTIFICACION",
        "TIPOID": "TIPO IDENTIFICACION",
        "CATEGORIA": "CATEGORÍA",
        "PLAN": "PLAN",
        "IDPLAN": "PLAN",
        "PL": "PLAN",
        "SEDE": "SEDE",
        "SED": "SEDE",
        "IDSEDE": "SEDE",
        "FINICIOCURSO": "F. INICIO CURSO",
        "FECINICIOCURSO": "F. INICIO CURSO",
        "DIASCONG": "DÍAS CONG.",
        "FECHAFINCURSOMASCONG": "FECHA FIN CURSO MÁS CONG.",
        "FECHAFINCURSOMASCONGEL": "FECHA FIN CURSO MÁS CONG.",
        "FASE": "FASE",
        "NIVEL": "NIVEL",
        "ESTADO": "ESTADO",
        "ESTADOEST": "ESTADO",
        "CONTACTOTITULAR": "CONTACTO TÍTULAR",
        "MOVILESTUDIANTE": "CONTACTO TÍTULAR",
        "TELEFONOTITULAR": "CONTACTO TÍTULAR",
        "FECHANAC": "FECHA NAC.",
        "FECHANACIMIENTO": "FECHA NAC.",
        "NOMBREPROGRAMA": "NOMBRE PROGRAMA",
        "CODPROGRAMA": "COD PROGRAMA",
        "CODIGOPROGRAMA": "COD PROGRAMA",
        "PROGRAMA": "NOMBRE PROGRAMA",
        "MODALIDAD": "MODALIDAD PREFERIDA",
        "MODALIDADPREFERIDA": "MODALIDAD PREFERIDA",
        "PREFERENCIAMODALIDAD": "MODALIDAD PREFERIDA",
        "PREFERENCIAMOD": "MODALIDAD PREFERIDA",
    }

    HEADER_PRIORITY = {
        "PLAN": {"PLAN": 1, "IDPLAN": 2, "PL": 3},
        "SEDE": {"SEDE": 1, "SED": 2, "IDSEDE": 3},
    }

    REQUIRED_COLUMNS = {"CÓDIGO USUARIO", "DOCUMENTO", "PLAN"}
    PROGRAM_COLUMNS = {"CATEGORÍA", "NOMBRE PROGRAMA", "COD PROGRAMA"}

    # ========================================================================
    # NORMALIZACIÓN DE CATEGORÍA → PROGRAMA
    # ========================================================================

    CATEGORIA_TO_PROGRAMA = {
        "ADULTOS": "Benglish",
        "B TEENS": "B teens",
    }

    # ========================================================================
    # NORMALIZACIÓN DE ESTADOS
    # ========================================================================

    ESTADO_MAPPING = {
        "ACTIVO": "active",
        "SUSPENDIDO": "inactive",
        "FINALIZADO": "graduated",
        "N/A": "inactive",
        "": "inactive",
    }

    # ========================================================================
    # MÉTODOS PRINCIPALES
    # ========================================================================

    def action_import(self):
        """Ejecuta la importación del archivo Excel"""
        self.ensure_one()

        if not openpyxl:
            raise UserError(
                _(
                    "La librería 'openpyxl' no está instalada.\n"
                    "Instálela con: pip install openpyxl"
                )
            )

        self.write({"state": "processing"})

        try:
            # Leer archivo Excel
            workbook = self._read_excel_file()
            sheet = workbook.active

            # Validar columnas
            self._validate_columns(sheet)

            # Procesar filas
            results = self._process_rows(sheet)

            # Actualizar estadísticas
            self.write(
                {
                    "state": "done",
                    "total_rows": results["total"],
                    "success_count": results["success"],
                    "error_count": results["errors"],
                    "skipped_count": results["skipped"],
                }
            )

            return self._show_results()

        except Exception as e:
            self.write({"state": "error"})
            self._log_error(0, "ERROR GENERAL", str(e))
            raise UserError(_("Error al procesar el archivo: %s") % str(e))

    def _read_excel_file(self):
        """Lee el archivo Excel desde el campo binario"""
        if not self.file_data:
            raise UserError(_("No se ha cargado ningún archivo"))

        file_content = base64.b64decode(self.file_data)
        return openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

    def _validate_columns(self, sheet):
        """Valida que las columnas del Excel coincidan con lo esperado"""
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        mapped_keys = set()
        for col in header_row:
            mapped = self._map_header_to_key(col)
            if mapped:
                mapped_keys.add(mapped)

        missing = [col for col in self.REQUIRED_COLUMNS if col not in mapped_keys]
        if not (self.PROGRAM_COLUMNS & mapped_keys):
            missing.append("CATEGORÍA / NOMBRE PROGRAMA / COD PROGRAMA")

        if missing:
            raise UserError(
                _(
                    "El archivo Excel no tiene las columnas esperadas.\n"
                    "Columnas faltantes: %s"
                )
                % ", ".join(missing)
            )

        _logger.info("✅ Validación de columnas exitosa")

    def _process_rows(self, sheet):
        """Procesa todas las filas del Excel"""
        results = {"total": 0, "success": 0, "errors": 0, "skipped": 0}

        # Obtener índices de columnas
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col_indices = {col: idx for idx, col in enumerate(header_row) if col}

        # Procesar cada fila (desde la fila 2)
        for row_num, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            results["total"] += 1

            try:
                # Extraer datos de la fila
                row_data = self._extract_row_data(row, col_indices)

                # Validar categoría (FILTRO CRÍTICO)
                if not self._has_program_info(row_data):
                    results["skipped"] += 1
                    self._log_info(
                        row_num,
                        row_data.get("CÓDIGO USUARIO", ""),
                        "Omitido: sin categoría o programa",
                    )
                    continue

                if row_data.get("CATEGORÍA") and not self._is_valid_categoria(
                    row_data.get("CATEGORÍA", "")
                ):
                    results["skipped"] += 1
                    self._log_info(
                        row_num,
                        row_data.get("CÓDIGO USUARIO", ""),
                        f"Omitido: categoría '{row_data.get('CATEGORÍA')}' no permitida",
                    )
                    continue

                # Importar el estudiante y matrícula
                import_result = self._import_student_and_enrollment(row_num, row_data)

                if import_result == "skipped":
                    results["skipped"] += 1
                    # _import_student_and_enrollment already logged the reason
                    continue

                results["success"] += 1
                self._log_success(
                    row_num,
                    row_data.get("CÓDIGO USUARIO", ""),
                    "Importado exitosamente",
                )

            except Exception as e:
                results["errors"] += 1
                error_msg = str(e)
                self._log_error(
                    row_num,
                    (
                        row_data.get("CÓDIGO USUARIO", "DESCONOCIDO")
                        if "row_data" in locals()
                        else "DESCONOCIDO"
                    ),
                    error_msg,
                )

                if not self.skip_errors:
                    raise
                else:
                    _logger.warning(f"Error en fila {row_num}: {error_msg}")
                    continue

        return results

    def _extract_row_data(self, row, col_indices):
        """Extrae los datos de una fila en un diccionario"""
        data = {}
        key_priority = {}
        for col_name, col_idx in col_indices.items():
            value = row[col_idx] if col_idx < len(row) else None

            # Convertir valores vacíos, guiones y ceros a None
            if value in (None, "", "-", 0, "0"):
                value = None
            elif isinstance(value, str):
                value = value.strip()
                if value in ("", "-", "0"):
                    value = None

            mapped_key = self._map_header_to_key(col_name)
            if not mapped_key:
                continue
            if value is None:
                continue
            source_norm = self._normalize_header_name(col_name)
            priority_map = self.HEADER_PRIORITY.get(mapped_key, {})
            new_priority = priority_map.get(source_norm, 50)
            current_priority = key_priority.get(mapped_key, 99)
            if mapped_key not in data or new_priority < current_priority:
                data[mapped_key] = value
                key_priority[mapped_key] = new_priority

        return data

    def _normalize_header_name(self, header):
        """Normaliza el nombre de columna para comparar con alias."""
        if not header:
            return ""
        header_norm = self._normalize_text(header)
        return re.sub(r"[^A-Z0-9]", "", header_norm)

    def _map_header_to_key(self, header):
        """Mapea un nombre de columna a la clave interna esperada."""
        normalized = self._normalize_header_name(header)
        return self.HEADER_ALIASES.get(normalized)

    # ========================================================================
    # VALIDACIONES Y NORMALIZACIONES
    # ========================================================================

    def _has_program_info(self, data):
        """Valida si la fila trae información mínima de programa/categoría."""
        return bool(
            (data or {}).get("CATEGORÍA")
            or (data or {}).get("NOMBRE PROGRAMA")
            or (data or {}).get("COD PROGRAMA")
        )

    def _is_valid_categoria(self, categoria):
        """Valida si la categoría está permitida"""
        if not categoria:
            return False

        categoria_norm = categoria.strip().upper()
        return categoria_norm in self.CATEGORIA_TO_PROGRAMA

    def _normalize_categoria(self, categoria):
        """Normaliza CATEGORÍA → Programa"""
        if not categoria:
            raise ValidationError(_("Categoría vacía"))

        categoria_norm = categoria.strip().upper()
        programa_name = self.CATEGORIA_TO_PROGRAMA.get(categoria_norm)

        if not programa_name:
            raise ValidationError(_("Categoría '%s' no válida") % categoria)

        return programa_name

    def _resolve_program(self, data, row_num=None):
        """Resuelve el programa desde categoría, nombre o código."""
        categoria = data.get("CATEGORÍA")
        if categoria:
            programa_name = self._normalize_categoria(categoria)
            programa = self.env["benglish.program"].search(
                [("name", "=", programa_name)], limit=1
            )
            if programa:
                return programa

        program_code = data.get("COD PROGRAMA")
        program_name = data.get("NOMBRE PROGRAMA")
        Program = self.env["benglish.program"]

        if program_code:
            programa = Program.search(
                [("code", "=ilike", str(program_code).strip())], limit=1
            )
            if programa:
                return programa

        if program_name:
            program_name_str = str(program_name).strip()
            programa = Program.search(
                [("name", "=ilike", program_name_str)], limit=1
            )
            if programa:
                return programa

            program_name_norm = self._normalize_text(program_name_str)
            for prog in Program.search([]):
                if self._normalize_text(prog.name) == program_name_norm:
                    return prog

        if row_num:
            codigo = data.get("CÓDIGO USUARIO", "DESCONOCIDO")
            self._log_info(
                row_num,
                codigo,
                "Programa no encontrado: revise CATEGORÍA/NOMBRE PROGRAMA/COD PROGRAMA",
            )

        return None

    def _normalize_plan(self, plan_excel, programa=None):
        """
        Normaliza el plan según la regla:
        plan_excel = UPPER(TRIM(valor))
        plan_sistema = "PLAN " + plan_excel

        CRÍTICO: Filtra por programa para evitar confusión entre planes con mismo nombre
        (ej: Plan PREMIUM de Benglish vs Plan PREMIUM de B-Teens)
        """
        if not plan_excel:
            return None

        plan_norm = plan_excel.strip().upper()
        plan_sistema = f"PLAN {plan_norm}"

        # Intentos de búsqueda tolerante (de más específico a más general)
        Plan = self.env["benglish.plan"]

        # Construir dominio base con programa si está disponible
        base_domain = [("program_id", "=", programa.id)] if programa else []

        _logger.info(
            f"🔍 Buscando plan '{plan_excel}' para programa {programa.name if programa else 'None'}"
        )

        # Intento 1: Buscar por "PLAN X" + programa
        plan = Plan.search(base_domain + [("name", "=ilike", plan_sistema)], limit=1)
        if plan:
            _logger.info(
                f"✅ Plan encontrado (nombre sistema): {plan.name} (ID: {plan.id})"
            )
            return plan

        # Intento 2: Buscar por nombre sin "PLAN" + programa
        plan = Plan.search(base_domain + [("name", "=ilike", plan_norm)], limit=1)
        if plan:
            _logger.info(
                f"✅ Plan encontrado (nombre simple): {plan.name} (ID: {plan.id})"
            )
            return plan

        # Intento 3: Buscar por código + programa
        plan = Plan.search(base_domain + [("code", "=ilike", plan_norm)], limit=1)
        if plan:
            _logger.info(
                f"✅ Plan encontrado (por código): {plan.name} (ID: {plan.id})"
            )
            return plan

        _logger.warning(
            f"❌ Plan '{plan_excel}' no encontrado en el sistema para el programa {programa.name if programa else '(sin programa)'}. "
            f"Esperados: '{plan_sistema}' o '{plan_norm}'. Fila será ignorada."
        )

        # Mostrar planes disponibles para este programa
        if programa:
            available_plans = Plan.search([("program_id", "=", programa.id)])
            _logger.info(
                f"🔍 Planes disponibles para {programa.name}: {[p.name for p in available_plans]}"
            )

        return None

    def _normalize_fase(self, fase_excel, programa=None):
        """Normaliza la fase académica.

        Comportamiento modificado: si la fase está vacía o no está entre las
        permitidas (BASIC/INTERMEDIATE/ADVANCED), se devuelve None en vez de
        levantar un error — la fila seguirá importándose y la fase será
        ignorada.
        """
        _logger.info(
            f"🔍 _normalize_fase llamada con: fase_excel='{fase_excel}', programa={programa.name if programa else 'None'}"
        )

        if not fase_excel:
            _logger.warning("⚠️ Fase vacía en Excel - será ignorada")
            return None

        fase_norm = fase_excel.strip().upper()
        _logger.info(f"🔍 Fase normalizada: '{fase_norm}'")

        valid_fases = ["BASIC", "INTERMEDIATE", "ADVANCED"]
        if fase_norm not in valid_fases:
            _logger.warning(
                f"❌ Fase no permitida en Excel: '{fase_excel}' (normalizada: '{fase_norm}') — será ignorada"
            )
            return None

        # Buscar la fase en el sistema (case-insensitive)
        domain = [("name", "=ilike", fase_norm)]
        # Si nos pasó el programa, filtrar por programa para distinguir BE vs BT
        if programa and getattr(programa, "id", False):
            domain.append(("program_id", "=", programa.id))
            _logger.info(f"🔍 Buscando fase con dominio: {domain}")

        fase = self.env["benglish.phase"].search(domain, limit=1)

        if not fase:
            prog_name = (
                programa.name
                if programa and getattr(programa, "name", False)
                else "(<sin programa>)"
            )
            _logger.warning(
                f"❌ Fase permitida ('{fase_norm}') no encontrada para el programa {prog_name} — será ignorada"
            )
            # Mostrar fases disponibles para este programa
            if programa:
                available_phases = self.env["benglish.phase"].search(
                    [("program_id", "=", programa.id)]
                )
                _logger.info(
                    f"🔍 Fases disponibles para {programa.name}: {[p.name for p in available_phases]}"
                )
            return None

        _logger.info(f"✅ Fase encontrada: {fase.name} (ID: {fase.id})")
        return fase

    def _parse_nivel(self, nivel_excel):
        """
        Extrae el número mayor del campo NIVEL
        Ejemplos:
        - "1 - 2" → 2
        - "11 - 12" → 12
        - "23 - 24" → 24
        """
        _logger.info(f"🔍 _parse_nivel llamada con: nivel_excel='{nivel_excel}'")

        # Si el campo NIVEL está vacío, no debe romper la importación.
        # Retornamos None para indicar que no hay información de nivel.
        if not nivel_excel:
            _logger.warning("⚠️ Nivel vacío en Excel")
            return None

        # Extraer todos los números
        numeros = re.findall(r"\d+", str(nivel_excel))
        _logger.info(f"🔍 Números extraídos de '{nivel_excel}': {numeros}")

        if not numeros:
            # No se pudo extraer números: tratar como información ausente
            _logger.warning(f"⚠️ No se pudieron extraer números de '{nivel_excel}'")
            return None

        # Tomar el mayor
        nivel_num = max(int(n) for n in numeros)
        _logger.info(f"✅ Nivel extraído: {nivel_num}")
        return nivel_num

    def _normalize_estado(self, estado_excel):
        """Normaliza el estado académico"""
        estado_norm = (estado_excel or "").strip().upper()
        return self.ESTADO_MAPPING.get(estado_norm, "inactive")

    def _parse_fecha(self, fecha_value):
        """Parsea fecha con tolerancia a diferentes formatos"""
        if not fecha_value:
            return None

        if isinstance(fecha_value, datetime):
            return fecha_value.date()

        if isinstance(fecha_value, str):
            # Intentar parsear diferentes formatos
            formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"]
            for fmt in formats:
                try:
                    return datetime.strptime(fecha_value.strip(), fmt).date()
                except ValueError:
                    continue

        # Si no se puede parsear, retornar None (no debe romper la importación)
        return None

    def _parse_telefono(self, telefono_value):
        """Valida y limpia el número telefónico"""
        if not telefono_value:
            return None

        telefono_str = str(telefono_value).strip()

        # Ignorar valores inválidos
        if telefono_str in ("-", "1", "0"):
            return None

        # Limpiar caracteres no numéricos
        telefono_clean = re.sub(r"[^\d+]", "", telefono_str)

        if len(telefono_clean) < 7:  # Mínimo 7 dígitos
            return None

        return telefono_clean

    def _normalize_documento(self, doc_value):
        """
        Normaliza el documento de identidad:
        - Elimina .0 de Excel
        - Elimina espacios, guiones, puntos
        - Conserva ceros a la izquierda
        - Convierte notación científica a número normal
        """
        if not doc_value:
            return None

        # Si es número (float/int), convertir a int primero para eliminar .0
        if isinstance(doc_value, (int, float)):
            doc_value = int(doc_value)

        # Convertir a string y limpiar
        doc_str = str(doc_value).strip()

        # Eliminar caracteres no numéricos
        doc_str = re.sub(r"[^\d]", "", doc_str)

        if not doc_str:
            return None

        return doc_str

    def _get_identification_type(self, tipo_code):
        """Obtiene el tipo de identificacion por codigo interno (cc/ti/ce/ppp)."""
        IdentificationType = self.env["l10n_latam.identification.type"]

        # Usar cache a nivel de módulo para evitar setear atributos en el recordset
        # (en ciertos contextos esto puede lanzar AttributeError).
        global _ID_TYPE_CACHE
        if tipo_code in _ID_TYPE_CACHE:
            cached_id = _ID_TYPE_CACHE.get(tipo_code)
            # Devolver un recordset en el env actual (si existe el id en cache)
            if cached_id:
                return IdentificationType.browse(cached_id)
            # Si el cache tiene None, devolvemos un recordset vacío
            return IdentificationType.browse()

        id_type = IdentificationType.browse()
        # Buscar por múltiples patrones comunes para cada tipo
        if tipo_code == "cc":
            patterns = [
                "cedula de ciudadan",
                "cedula ciudadan",
                "cedula",
                "cédula",
                "cedula de identidad",
                "cc",
                "c.c",
            ]
            for p in patterns:
                id_type = IdentificationType.search([("name", "ilike", p)], limit=1)
                if id_type:
                    break
            # Intentar por código si existe campo 'code'
            if not id_type:
                try:
                    id_type = IdentificationType.search([("code", "ilike", "CC")], limit=1)
                except Exception:
                    pass
        elif tipo_code == "ti":
            patterns = [
                "tarjeta de identidad",
                "tarjeta identidad",
                "tarjeta",
                "ti",
                "t.i",
            ]
            for p in patterns:
                id_type = IdentificationType.search([("name", "ilike", p)], limit=1)
                if id_type:
                    break
            if not id_type:
                try:
                    id_type = IdentificationType.search([("code", "ilike", "TI")], limit=1)
                except Exception:
                    pass
        elif tipo_code == "ce":
            patterns = [
                "cedula de extranjer",
                "cedula extranjer",
                "cedula de extranjeria",
                "extranjeria",
                "ce",
            ]
            for p in patterns:
                id_type = IdentificationType.search([("name", "ilike", p)], limit=1)
                if id_type:
                    break
            if not id_type:
                try:
                    id_type = IdentificationType.search([("code", "ilike", "CE")], limit=1)
                except Exception:
                    pass
        elif tipo_code == "ppp":
            for pattern in (
                "permiso para trabajar",
                "permiso por proteccion temporal",
                "permiso especial de permanencia",
                "permiso de permanencia",
                "permiso temporal",
                "ppt",
                "pep",
                "ppp",
            ):
                id_type = IdentificationType.search([("name", "ilike", pattern)], limit=1)
                if id_type:
                    break

        # Guardar solo el id en el cache y devolver un recordset en el env actual
        try:
            stored_id = id_type.id if id_type else None
            _ID_TYPE_CACHE[tipo_code] = stored_id
        except Exception:
            # Fallback: no bloquear por errores de cache
            _ID_TYPE_CACHE[tipo_code] = None

        return IdentificationType.browse(_ID_TYPE_CACHE.get(tipo_code))

    def _get_document_type_value(self, data):
        """Obtiene el tipo de documento desde la fila, si viene en el Excel."""
        if not data:
            return None
        candidates = {
            "TIPODOCUMENTO",
            "TIPODOC",
            "TIPOIDENTIFICACION",
            "TIPOID",
            "DOCUMENTOTIPO",
            "TIPO",
        }
        for key, value in data.items():
            key_norm = self._normalize_text(key)
            key_compact = re.sub(r"[^A-Z0-9]", "", key_norm)
            if key_compact in candidates:
                return value
        return None

    def _normalize_doc_type_code(self, doc_type_value):
        """Normaliza el tipo de documento del Excel (CC/TI/CE/PPP)."""
        if not doc_type_value:
            return None
        doc_text = self._normalize_text(doc_type_value)
        doc_compact = re.sub(r"[^A-Z]", "", doc_text)

        if doc_compact in ("CC", "CEDULACIUDADANIA"):
            return "cc"
        if doc_compact in ("TI", "TARJETAIDENTIDAD"):
            return "ti"
        if doc_compact in ("CE", "CEDULAEXTRANJERIA"):
            return "ce"
        if doc_compact in ("PPP", "PPT", "PEP"):
            return "ppp"

        if "EXTRANJERIA" in doc_text:
            return "ce"
        if "TARJETA" in doc_text:
            return "ti"
        if "PERMISO" in doc_text:
            return "ppp"
        if "CEDULA" in doc_text:
            return "cc"

        return None

    def _normalize_delivery_mode(self, mode_value):
        """Normaliza el valor de modalidad del Excel a los valores internos.

        Convierte variantes en español/inglés a: 'presential'|'virtual'|'hybrid'.
        """
        if not mode_value:
            return None

        txt = self._normalize_text(mode_value)
        # Eliminar espacios y caracteres no alfabéticos
        compact = re.sub(r"[^A-Z]", "", txt)

        # Mapas comunes
        presential_keys = ["PRESENCIAL", "PRESENCE", "PRES", "P"]
        virtual_keys = ["VIRTUAL", "VIRT", "ONLINE", "REMOTO", "REMOTE", "V"]
        hybrid_keys = ["HIBRIDO", "HÍBRIDO", "HYBRID", "H"]

        if any(k in txt for k in ("PRESENCIAL", "PRESENCE")) or compact in ("PRES", "P"):
            return "presential"
        if any(k in txt for k in ("VIRTUAL", "ONLINE", "REMOTO", "REMOTE")) or compact in ("VIRT", "V"):
            return "virtual"
        if any(k in txt for k in ("HIBRIDO", "HYBRID")) or compact == "H":
            return "hybrid"

        # Fallback: si el texto contiene palabras clave en inglés
        if "PRES" in compact:
            return "presential"
        if "VIRT" in compact or "ONL" in compact:
            return "virtual"
        if "HIB" in compact or "HYB" in compact:
            return "hybrid"

        return None

    def _infer_id_type_from_documento(
        self, doc_value, birth_date=None, doc_type_value=None
    ):
        """Infiere el tipo de documento usando texto crudo y edad como fallback."""
        tipo_code = self._normalize_doc_type_code(doc_type_value)
        if tipo_code:
            id_type = self._get_identification_type(tipo_code)
            if not id_type:
                _logger.warning(
                    "Tipo de documento %s no encontrado en el sistema.", tipo_code
                )
            return id_type

        doc_text = self._normalize_text(doc_value) if doc_value else ""
        if doc_text:
            doc_compact = re.sub(r"[^A-Z]", "", doc_text)
            if (
                re.search(r"\bT\s*\.?\s*I\b", doc_text)
                or doc_compact.startswith("TI")
                or "TARJETA" in doc_text
            ):
                id_type = self._get_identification_type("ti")
                if not id_type:
                    _logger.warning("Tipo de documento TI no encontrado en el sistema.")
                return id_type
            if (
                re.search(r"\bC\s*\.?\s*C\b", doc_text)
                or doc_compact.startswith("CC")
                or "CEDULA" in doc_text
            ):
                id_type = self._get_identification_type("cc")
                if not id_type:
                    _logger.warning("Tipo de documento CC no encontrado en el sistema.")
                return id_type

        if birth_date:
            today = fields.Date.context_today(self)
            age = (
                today.year
                - birth_date.year
                - ((today.month, today.day) < (birth_date.month, birth_date.day))
            )
            tipo_code = "cc" if age >= 18 else "ti"
            id_type = self._get_identification_type(tipo_code)
            if not id_type:
                _logger.warning(
                    "Tipo de documento %s no encontrado en el sistema.", tipo_code
                )
            return id_type

        return None

    def _normalize_text(self, text):
        """Normaliza texto: sin tildes, minúsculas, sin espacios extras"""
        if not text:
            return ""
        import unicodedata

        # Eliminar tildes/acentos
        text = unicodedata.normalize("NFD", str(text))
        text = "".join(char for char in text if unicodedata.category(char) != "Mn")
        # Convertir a mayúsculas y quitar espacios extras
        return text.upper().strip()

    def _find_campus(self, sede_name):
        """Busca la sede en el sistema con búsqueda MUY flexible (sin acentos, case-insensitive, tolerante)"""
        if not sede_name:
            return None

        sede_name_normalized = self._normalize_text(sede_name)

        if not sede_name_normalized:
            return None

        _logger.info(
            f"🔍 Buscando sede: '{sede_name}' (normalizado: '{sede_name_normalized}')"
        )

        # Buscar en TODAS las sedes y comparar sin tildes/mayúsculas
        all_campuses = self.env["benglish.campus"].search([])

        for campus in all_campuses:
            campus_name_normalized = self._normalize_text(campus.name)
            campus_city_normalized = (
                self._normalize_text(campus.city_name) if campus.city_name else ""
            )
            campus_code_normalized = (
                self._normalize_text(campus.code) if campus.code else ""
            )

            # Intento 1: Coincidencia exacta del nombre normalizado
            if sede_name_normalized == campus_name_normalized:
                _logger.info(
                    f"✅ Sede encontrada (nombre exacto): '{sede_name}' -> {campus.name} (ciudad: {campus.city_name})"
                )
                return campus

            # Intento 2: Coincidencia exacta del código normalizado
            if sede_name_normalized == campus_code_normalized:
                _logger.info(
                    f"✅ Sede encontrada (código exacto): '{sede_name}' -> {campus.name} (ciudad: {campus.city_name})"
                )
                return campus

            # Intento 3: Coincidencia exacta con la ciudad
            if sede_name_normalized == campus_city_normalized:
                _logger.info(
                    f"✅ Sede encontrada (por ciudad): '{sede_name}' -> {campus.name} (ciudad: {campus.city_name})"
                )
                return campus

            # Intento 4: El nombre del Excel está contenido en el nombre de la sede
            if sede_name_normalized in campus_name_normalized:
                _logger.info(
                    f"✅ Sede encontrada (contiene en nombre): '{sede_name}' -> {campus.name} (ciudad: {campus.city_name})"
                )
                return campus

            # Intento 5: El nombre de la sede está contenido en el Excel
            if campus_name_normalized in sede_name_normalized:
                _logger.info(
                    f"✅ Sede encontrada (sede contenida en Excel): '{sede_name}' -> {campus.name} (ciudad: {campus.city_name})"
                )
                return campus

        # Si no encontró nada
        _logger.warning(
            f"⚠️ No se encontró sede para: '{sede_name}' (normalizado: '{sede_name_normalized}')"
        )
        return None

        # Si no se encuentra, listar sedes disponibles para ayudar
        available = ", ".join([f"{c.code}-{c.name}" for c in all_campuses[:20]])
        _logger.warning(
            f"❌ Sede '{sede_name_clean}' NO encontrada. Sedes disponibles: {available}"
        )

        return None

    def _get_default_freeze_reason(self):
        """Obtiene (o crea) un motivo de congelamiento por defecto para imports.
        Reglas:
        - Buscar un motivo activo cuyo nombre o código contenga 'import'
        - Si no existe, usar el primer motivo activo y no especial
        - Si aún no existe ninguno, crear un motivo 'Importado desde Excel'
        """
        Reason = self.env["benglish.freeze.reason"]

        # Buscar motivo que indique importación
        reason = Reason.search(
            [("active", "=", True), ("name", "ilike", "import")], limit=1
        )
        if not reason:
            reason = Reason.search(
                [("active", "=", True), ("code", "ilike", "import")], limit=1
            )

        # Si no, elegir un motivo activo y no especial
        if not reason:
            reason = Reason.search(
                [("active", "=", True), ("es_especial", "=", False)], limit=1
            )

        # Si sigue sin motivos, crear uno genérico
        if not reason:
            reason = Reason.create(
                {
                    "name": "Importado desde Excel",
                    "code": "IMPORTADO",
                    "active": True,
                    "es_especial": False,
                    "sequence": 100,
                }
            )

        return reason

    # ========================================================================
    # IMPORTACIÓN DE ESTUDIANTE Y MATRÍCULA
    # ========================================================================

    def _import_student_and_enrollment(self, row_num, data):
        """
        Importa un estudiante y su matrícula siguiendo el orden:
        1. Validar categoría
        2. Normalizar programa
        3. Normalizar plan
        4. Crear/actualizar estudiante (incluye sede principal, ciudad y país)
        5. Crear matrícula
        7. Asignar fase
        8. Procesar niveles → asistencia histórica
        9. Procesar congelamientos
        10. Aplicar estado académico final
        """

        # 1-3. Validaciones y normalizaciones
        programa = self._resolve_program(data, row_num=row_num)

        if not programa:
            raise ValidationError(
                _("Programa no encontrado en el sistema")
            )

        plan = self._normalize_plan(data.get("PLAN"), programa)  # ← PASAR programa aquí
        if not plan:
            # Plan no existe — omitir la fila
            codigo = data.get("CÓDIGO USUARIO", "DESCONOCIDO")
            self._log_info(
                row_num,
                codigo,
                f"Omitido: Plan '{data.get('PLAN')}' no existe en el sistema para programa {programa.name}",
            )
            return "skipped"
        fase_excel = data.get("FASE")
        _logger.info(
            f"🔍 Procesando FASE desde Excel: '{fase_excel}' para programa {programa.name}"
        )

        fase = self._normalize_fase(fase_excel, programa)
        if not fase:
            # Registrar en el log que la fase fue ignorada (no bloquear la fila)
            codigo = data.get("CÓDIGO USUARIO", "DESCONOCIDO")
            fase_val = fase_excel or "(vacía)"
            self._log_info(
                row_num,
                codigo,
                f"Fase '{fase_val}' no permitida o no encontrada — ignorada",
            )
            _logger.warning(f"⚠️ Fase NO asignada para estudiante {codigo}")
        else:
            _logger.info(f"✅ Fase lista para asignar: {fase.name} (ID: {fase.id})")

        # 4. Crear o actualizar estudiante (incluye sede principal, ciudad y país)
        student = self._create_or_update_student(data, row_num=row_num)
        _logger.info(f"✅ Estudiante procesado: {student.code} - {student.name}")

        # 5. Crear matrícula
        _logger.info(
            f"🔍 Creando matrícula para {student.code} con fase: {fase.name if fase else 'None'}"
        )
        enrollment = self._create_enrollment(
            student, programa, plan, fase, data, row_num
        )

        # Si la creación fue omitida por duplicado, omitir la fila completa
        if enrollment == "skipped":
            return "skipped"

        # 6. La fase y nivel ya se asignaron en la matrícula

        # 7. Procesar niveles → asistencia histórica
        self._process_asistencia_historica(
            enrollment, fase, data.get("NIVEL"), data.get("ESTADO")
        )

        # 8. Procesar congelamientos
        self._process_congelamientos(student, enrollment, data)

        # 9. Aplicar estado académico final
        self._apply_final_state(student, enrollment, data.get("ESTADO"))

        _logger.info(f"✅ Importado: {student.code} - {student.name}")

    def _is_valid_email(self, email):
        """Validación ligera de formato de email: devuelve True si parece válido."""
        if not email:
            return False
        # Regex simple: algo@algo.dominio (no pretende cubrir todos los casos RFC)
        pattern = r"^[^@\s]+@[^@\s]+\.[^@\s]+$"
        return re.match(pattern, str(email).strip()) is not None

    def _create_or_update_student(self, data, row_num=None):
        """Crea o actualiza un estudiante según CÓDIGO USUARIO"""
        codigo = data.get("CÓDIGO USUARIO")

        if not codigo:
            raise ValidationError(_("CÓDIGO USUARIO vacío"))

        # Buscar estudiante existente por código
        student = self.env["benglish.student"].search([("code", "=", codigo)], limit=1)

        # Buscar sede principal primero
        campus = self._find_campus(data.get("SEDE"))

        # Obtener país Colombia
        country_colombia = self.env["res.country"].search(
            [("code", "=", "CO")], limit=1
        )

        # Preparar valores
        # Priorizar Primer Nombre y Primer Apellido. Si faltan, usar '-'
        first_name = data.get("PRIMER NOMBRE") or "-"
        first_last_name = data.get("PRIMER APELLIDO") or "-"

        # Normalizar documento
        documento_normalizado = self._normalize_documento(data.get("DOCUMENTO"))
        birth_date = self._parse_fecha(data.get("FECHA NAC."))
        doc_type_value = self._get_document_type_value(data)
        id_type = self._infer_id_type_from_documento(
            data.get("DOCUMENTO"), birth_date, doc_type_value
        )

        # Manejar email: si el email está presente pero es inválido, OMITIRLO
        email_val = data.get("EMAIL")
        if email_val and isinstance(email_val, str):
            # Normalizar y aceptar prefijos 'mailto:' (p. ej. 'mailto:Maria@gmail.com')
            email_str = email_val.strip()
            if email_str.lower().startswith("mailto:"):
                # Quitar el prefijo y continuar con la dirección real
                email_str = email_str.split(":", 1)[1].strip()
                email_val = email_str

            if not self._is_valid_email(email_val):
                # Registrar en el log y no incluir el email en los valores
                if row_num:
                    codigo_log = data.get("CÓDIGO USUARIO", "DESCONOCIDO")
                    self._log_info(
                        row_num,
                        codigo_log,
                        f"Email inválido omitido: '{email_val}'",
                    )
                _logger.warning(
                    f"Email inválido omitido en fila {row_num}: {email_val}"
                )
                email_val = None

        values = {
            "code": codigo,
            "first_name": first_name,
            "second_name": data.get("SEGUNDO NOMBRE"),
            "first_last_name": first_last_name,
            "second_last_name": data.get("SEGUNDO APELLIDO"),
            "student_id_number": documento_normalizado,
            "mobile": self._parse_telefono(data.get("CONTACTO TÍTULAR")),
            "birth_date": birth_date,
            "enrollment_date": self._parse_fecha(data.get("F. INICIO CURSO"))
            or fields.Date.today(),
        }
        # Modalidad preferida desde Excel (si viene)
        modalidad_val = data.get("MODALIDAD PREFERIDA") or data.get("MODALIDAD")
        modalidad_norm = self._normalize_delivery_mode(modalidad_val)
        if modalidad_norm:
            values["preferred_delivery_mode"] = modalidad_norm
        if id_type:
            values["id_type_id"] = id_type.id

        # Asignar sede principal
        if campus:
            values["preferred_campus_id"] = campus.id
            _logger.info(f"✅ Sede asignada: {campus.name} (ID: {campus.id})")
            # Asignar ciudad desde sede
            if campus.city_name:
                values["city"] = campus.city_name
                _logger.info(f"✅ Ciudad asignada desde sede: {campus.city_name}")
            else:
                _logger.warning(f"⚠️ Sede {campus.name} no tiene city_name configurado")
        else:
            _logger.warning(f"⚠️ No se asignó sede para estudiante {codigo}")

        # Asignar país Colombia por defecto
        if country_colombia:
            values["country_id"] = country_colombia.id
            _logger.info(f"✅ País asignado: Colombia (ID: {country_colombia.id})")
        else:
            _logger.warning("⚠️ No se encontró el país Colombia en el sistema")

        # Añadir email solo si es válido/no nulo
        if email_val:
            values["email"] = email_val

        # Limpiar SOLO valores None (NO eliminar strings vacíos válidos)
        values = {k: v for k, v in values.items() if v is not None}

        if student and self.update_existing:
            # Actualizar existente
            student.write(values)
            _logger.info(f"Actualizado estudiante: {codigo}")
        elif not student:
            # Crear nuevo
            student = self.env["benglish.student"].create(values)
            _logger.info(f"Creado estudiante: {codigo}")
        else:
            _logger.info(f"Estudiante ya existe (sin actualizar): {codigo}")

        return student

    def _create_enrollment(self, student, programa, plan, fase, data, row_num=None):
        """Crea la matrícula del estudiante"""

        fecha_inicio = (
            self._parse_fecha(data.get("F. INICIO CURSO")) or fields.Date.today()
        )
        fecha_fin = self._parse_fecha(data.get("FECHA FIN CURSO MÁS CONG."))

        # Antes de crear, comprobar si ya existe una matrícula activa para el mismo estudiante y plan
        closed_states = ["finished", "cancelled", "withdrawn", "failed", "completed"]
        existing = self.env["benglish.enrollment"].search(
            [
                ("student_id", "=", student.id),
                ("plan_id", "=", plan.id),
                ("state", "not in", closed_states),
            ],
            limit=1,
        )

        if existing:
            # Registrar y omitir la creación: el usuario pidió "omitir" en caso de duplicados
            codigo = student.code or "DESCONOCIDO"
            mensaje = f"Omitido: matrícula existente activa para estudiante {codigo} en el plan '{plan.name}'"
            if row_num:
                self._log_info(row_num, codigo, mensaje)
            _logger.info(mensaje)
            return "skipped"

        values = {
            "student_id": student.id,
            "program_id": programa.id,
            "plan_id": plan.id,
        }

        # Establecer modalidad de la matrícula: preferir valor del Excel, si no usar preferencia del estudiante
        modalidad_val = (
            data.get("MODALIDAD PREFERIDA") or data.get("MODALIDAD") or None
        )
        modalidad_norm = self._normalize_delivery_mode(modalidad_val) or (
            student.preferred_delivery_mode if getattr(student, "preferred_delivery_mode", None) else None
        )
        if modalidad_norm:
            values["delivery_mode"] = modalidad_norm

        # Establecer fase si existe
        if fase:
            values["current_phase_id"] = fase.id
            _logger.info(f"✅ Fase asignada en matrícula: {fase.name} (ID: {fase.id})")
        else:
            _logger.warning(f"⚠️ No se asignó fase para matrícula de {student.code}")

        # Asignar nivel actual basado en fase y número de nivel
        nivel_id = None
        if fase and data.get("NIVEL"):
            unidad_final = self._parse_nivel(data.get("NIVEL"))
            _logger.info(
                f"🔍 Buscando nivel para fase {fase.name}, unidad final: {unidad_final}"
            )
            if unidad_final:
                # Buscar el nivel cuyo max_unit sea mayor o igual a la unidad final
                # y que sea el más pequeño (para no asignar un nivel muy avanzado)
                niveles = self.env["benglish.level"].search(
                    [
                        ("phase_id", "=", fase.id),
                        ("max_unit", ">=", unidad_final),
                    ],
                    order="max_unit, sequence",
                    limit=1,
                )

                if niveles:
                    nivel_id = niveles[0].id
                    values["current_level_id"] = nivel_id
                    _logger.info(
                        f"✅ Nivel asignado en matrícula: {niveles[0].name} (ID: {nivel_id}, max_unit: {niveles[0].max_unit})"
                    )
                else:
                    # Si no encuentra nivel con max_unit >= unidad_final,
                    # buscar el nivel más avanzado de la fase (probablemente está en ese nivel o más allá)
                    nivel_max = self.env["benglish.level"].search(
                        [("phase_id", "=", fase.id)],
                        order="max_unit desc, sequence desc",
                        limit=1,
                    )
                    if nivel_max:
                        nivel_id = nivel_max.id
                        values["current_level_id"] = nivel_id
                        _logger.info(
                            f"✅ Nivel asignado (nivel máximo): {nivel_max.name} (ID: {nivel_id}, max_unit: {nivel_max.max_unit})"
                        )
                    else:
                        _logger.warning(
                            f"⚠️ No se encontró ningún nivel para fase {fase.name}"
                        )
            else:
                _logger.warning(
                    f"⚠️ No se pudo parsear el nivel desde: {data.get('NIVEL')}"
                )
        elif not fase:
            _logger.warning(f"⚠️ No se puede asignar nivel porque no hay fase")
        elif not data.get("NIVEL"):
            _logger.warning(
                f"⚠️ No se puede asignar nivel porque el campo NIVEL está vacío"
            )

        # completar el resto de valores
        values.update(
            {
                "enrollment_date": fecha_inicio,
                "course_start_date": fecha_inicio,
                "course_end_date": fecha_fin,
                "categoria": data.get("CATEGORÍA"),
                "state": "enrolled",
            }
        )

        enrollment = self.env["benglish.enrollment"].create(values)

        _logger.info(f"✅ Creada matrícula {enrollment.code} para {student.code}")

        # IMPORTANTE: Actualizar program_id y plan_id en el estudiante
        # (estos campos son readonly y solo se actualizan desde matrícula)
        student_vals = {}
        if programa and not student.program_id:
            student_vals["program_id"] = programa.id
            _logger.info(f"✅ Actualizando program_id del estudiante: {programa.name}")
        if plan and not student.plan_id:
            student_vals["plan_id"] = plan.id
            _logger.info(f"✅ Actualizando plan_id del estudiante: {plan.name}")

        if student_vals:
            student.write(student_vals)
            # Forzar recálculo de campos computados
            student.invalidate_recordset(["active_enrollment_ids"])
            student._compute_current_academic_info()

        return enrollment

    def _process_asistencia_historica(
        self, enrollment, fase, nivel_excel, estado_excel
    ):
        """
        Marca asistencia histórica en enrollment_progress_ids
        según el campo NIVEL del Excel
        """

        # Parsear nivel
        unidad_final = self._parse_nivel(nivel_excel)

        _logger.info(f"Procesando asistencia histórica hasta unidad {unidad_final}")

        # Obtener todas las asignaturas del plan
        plan = enrollment.plan_id

        # Buscar asignaturas a través de: plan -> fases -> niveles -> asignaturas
        subjects = self.env["benglish.subject"].search(
            [("level_id.phase_id.id", "in", plan.phase_ids.ids)], order="sequence"
        )

        # Si el estado es FINALIZADO, marcar TODAS como asistidas
        estado_norm = (estado_excel or "").strip().upper()
        marcar_todas = estado_norm == "FINALIZADO"

        # Si no hay información de nivel y no está FINALIZADO, omitir procesamiento
        if unidad_final is None and not marcar_todas:
            _logger.info(
                f"NIVEL vacío o no parseable para enrollment {enrollment.id}; se omite asistencia histórica"
            )
            return

        # Crear registros de progreso
        for subject in subjects:
            # Verificar si ya existe el registro de progreso
            existing_progress = self.env["benglish.enrollment.progress"].search(
                [
                    ("enrollment_id", "=", enrollment.id),
                    ("subject_id", "=", subject.id),
                ],
                limit=1,
            )

            # Determinar si asistió
            if marcar_todas:
                attended = True
                state = "completed"
            else:
                attended = subject.sequence <= unidad_final
                state = "in_progress" if attended else "pending"

            if existing_progress:
                # Actualizar el registro existente
                existing_progress.write(
                    {
                        "state": state,
                        "start_date": (
                            enrollment.course_start_date if attended else None
                        ),
                    }
                )
                _logger.debug(
                    f"Actualizado progreso para {subject.name} (seq={subject.sequence}): state={state}"
                )
            else:
                # Crear nuevo registro de progreso
                self.env["benglish.enrollment.progress"].create(
                    {
                        "enrollment_id": enrollment.id,
                        "subject_id": subject.id,
                        "state": state,
                        "start_date": (
                            enrollment.course_start_date if attended else None
                        ),
                    }
                )
                _logger.debug(
                    f"Creado progreso para {subject.name} (seq={subject.sequence}): state={state}"
                )

    def _process_congelamientos(self, student, enrollment, data):
        """
        Procesa congelamientos según DÍAS CONG.
        """

        dias_cong = data.get("DÍAS CONG.")

        if not dias_cong or dias_cong == 0:
            return  # No hay congelamiento

        try:
            dias_cong = int(dias_cong)
        except (ValueError, TypeError):
            _logger.warning(f"Días de congelamiento inválido: {dias_cong}")
            return

        if dias_cong <= 0:
            return

        # Crear registro de congelamiento
        fecha_inicio = (
            self._parse_fecha(data.get("F. INICIO CURSO")) or fields.Date.today()
        )
        fecha_fin = fecha_inicio + timedelta(days=dias_cong)

        freeze_values = {
            "student_id": student.id,
            "enrollment_id": enrollment.id,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "estado": "aprobado",  # Ya está aprobado (dato histórico)
            "es_especial": False,
            "motivo_especial": "Congelamiento importado desde Excel",
        }

        # Asegurar motivo de congelamiento requerido
        try:
            reason = self._get_default_freeze_reason()
            if reason:
                freeze_values["freeze_reason_id"] = reason.id
        except Exception:
            # No bloquear la importación solo por problemas al obtener/crear motivo
            _logger.exception(
                "No se pudo obtener/crear motivo de congelamiento por defecto"
            )

        freeze = self.env["benglish.student.freeze.period"].create(freeze_values)

        _logger.info(f"✅ Creado congelamiento de {dias_cong} días para {student.code}")

    def _apply_final_state(self, student, enrollment, estado_excel):
        """
        Aplica el estado académico final según campo ESTADO
        """

        estado = self._normalize_estado(estado_excel)

        # Actualizar estado del estudiante
        student.write({"state": estado})

        # Actualizar estado de la matrícula
        # Estados válidos: draft, pending, enrolled, active, in_progress,
        #                  suspended, completed, failed, finished, homologated, withdrawn, cancelled
        enrollment_state = "enrolled"
        if estado == "graduated":
            enrollment_state = "finished"
        elif estado == "inactive":
            enrollment_state = "suspended"

        enrollment.write({"state": enrollment_state})

        _logger.info(f"Estado final: estudiante={estado}, matrícula={enrollment_state}")

    # ========================================================================
    # LOGGING
    # ========================================================================

    def _log_success(self, row_num, student_code, message):
        """Registra un éxito en el log"""
        self.env["benglish.student.enrollment.import.log"].create(
            {
                "wizard_id": self.id,
                "row_number": row_num,
                "student_code": student_code,
                "message": message,
                "log_type": "success",
            }
        )

    def _log_error(self, row_num, student_code, message):
        """Registra un error en el log"""
        self.env["benglish.student.enrollment.import.log"].create(
            {
                "wizard_id": self.id,
                "row_number": row_num,
                "student_code": student_code,
                "message": message,
                "log_type": "error",
            }
        )

    def _log_info(self, row_num, student_code, message):
        """Registra información en el log"""
        self.env["benglish.student.enrollment.import.log"].create(
            {
                "wizard_id": self.id,
                "row_number": row_num,
                "student_code": student_code,
                "message": message,
                "log_type": "info",
            }
        )

    def _show_results(self):
        """Muestra los resultados de la importación"""
        return {
            "type": "ir.actions.act_window",
            "name": "Resultados de Importación",
            "res_model": self._name,
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
        }


class StudentEnrollmentImportLog(models.TransientModel):
    """Log de importación de estudiantes y matrículas"""

    _name = "benglish.student.enrollment.import.log"
    _description = "Log de Importación de Estudiantes"
    _order = "row_number"

    wizard_id = fields.Many2one(
        "benglish.student.enrollment.import.wizard",
        string="Wizard",
        required=True,
        ondelete="cascade",
    )

    row_number = fields.Integer(string="Fila")
    student_code = fields.Char(string="Código Estudiante")
    message = fields.Text(string="Mensaje")
    log_type = fields.Selection(
        [("success", "Éxito"), ("error", "Error"), ("info", "Información")],
        string="Tipo",
        default="info",
    )
