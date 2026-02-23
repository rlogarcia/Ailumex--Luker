# -*- coding: utf-8 -*-

import base64
import hashlib
import io
import logging
import mimetypes
import re
import unicodedata
from datetime import date, datetime

from odoo import api, fields, models, _
from odoo.exceptions import UserError
from odoo.tools import email_normalize

try:
    import openpyxl
except ImportError:  # pragma: no cover - handled in runtime
    openpyxl = None


_logger = logging.getLogger(__name__)


class StudentImportBatch(models.Model):
    _name = "benglish.student.import.batch"
    _description = "Batch de importación de estudiantes"
    _inherit = ["mail.thread", "mail.activity.mixin"]
    _order = "create_date desc, id desc"

    REQUIRED_COLUMNS = [
        "documento_identidad",
        "primer_nombre",
        "primer_apellido",
        "email",
        "telefono",
    ]
    OPTIONAL_COLUMNS = [
        "segundo_nombre",
        "segundo_apellido",
        "celular",
        "fecha_nacimiento",
        "genero",
        "codigo",
        "direccion",
        "ciudad",
        "pais",
        "programa",
        "nombre_programa",
        "plan",
        "plan_comercial",
        "sede",
        "fase",
        "nivel",
        "unidad",
        "modalidad",
        "categoria",
        "fecha_inicio_curso",
        "fecha_fin_curso",
        "fecha_maxima_congelamiento",
        "dias_curso",
        "contacto_titular",
        "estado_academico",
        "nombre_completo",  # Por compatibilidad
        # Campos adicionales del Excel de ejemplo
        "tipo_identificacion",
        "codigo_est",
        "cod_programa",
        "id_plan",
        "pl",
        "id_sede",
        "estado_est",
    ]
    COLUMN_ALIASES = {
        # Documento
        "documento": "documento_identidad",
        "documento_identidad": "documento_identidad",
        "documentoidentidad": "documento_identidad",
        "id": "documento_identidad",
        "identificacion": "documento_identidad",
        # Tipo de identificación
        "tipo_identificacion": "tipo_identificacion",
        "tipoidentificacion": "tipo_identificacion",
        "tipo_documento": "tipo_identificacion",
        # Nombres desagregados
        "primer_nombre": "primer_nombre",
        "primernombre": "primer_nombre",
        "nombre1": "primer_nombre",
        "segundo_nombre": "segundo_nombre",
        "segundonombre": "segundo_nombre",
        "nombre2": "segundo_nombre",
        "primer_apellido": "primer_apellido",
        "primerapellido": "primer_apellido",
        "pirmer_apellido": "primer_apellido",  # Typo común
        "apellido1": "primer_apellido",
        "segundo_apellido": "segundo_apellido",
        "segundoapellido": "segundo_apellido",
        "apellido2": "segundo_apellido",
        # Nombres completos (por compatibilidad)
        "nombres": "primer_nombre",  # Mapeo legacy
        "nombre": "primer_nombre",
        "apellidos": "primer_apellido",  # Mapeo legacy
        "apellido": "primer_apellido",
        "nombre_completo": "nombre_completo",
        "nombrecompleto": "nombre_completo",
        # Contacto
        "email": "email",
        "correo": "email",
        "correo_electronico": "email",
        "email_est": "email",
        "telefono": "telefono",
        "telefono_principal": "telefono",
        "celular": "celular",
        "movil": "celular",
        "movil_estudiante": "celular",
        # Datos personales
        "fecha_nacimiento": "fecha_nacimiento",
        "nacimiento": "fecha_nacimiento",
        "genero": "genero",
        "sexo": "genero",
        "codigo": "codigo",
        "codigo_estudiante": "codigo",
        "codigo_est": "codigo",
        # Dirección
        "direccion": "direccion",
        "direccion_residencia": "direccion",
        "ciudad": "ciudad",
        "pais": "pais",
        # Académico - Programa
        "programa": "programa",
        "nombre_programa": "programa",
        "cod_programa": "programa",
        # Académico - Plan Comercial (principal)
        "plan": "plan_comercial",
        "plan_comercial": "plan_comercial",
        "plancomercial": "plan_comercial",
        "id_plan": "plan_comercial",
        "pl": "plan_comercial",
        # Sede
        "sede": "sede",
        "id_sede": "sede",
        "sed": "sede",
        # Fase y Nivel
        "fase": "fase",
        "nivel": "nivel",
        "unidad": "nivel",
        # Modalidad
        "modalidad": "modalidad",
        # Categoría
        "categoria": "categoria",
        # Fechas del curso/matrícula
        "fecha_inicio_curso": "fecha_inicio_curso",
        "fecha_inicio": "fecha_inicio_curso",
        "fecha_fin_curso": "fecha_fin_curso",
        "fecha_fin": "fecha_fin_curso",
        "fecha_maxima_congelamiento": "fecha_maxima_congelamiento",
        "fecha_max_congelamiento": "fecha_maxima_congelamiento",
        "dias_curso": "dias_curso",
        "dias_del_curso": "dias_curso",
        "duracion_curso": "dias_curso",
        # Titular
        "contacto_titular": "contacto_titular",
        "titular": "contacto_titular",
        "nombre_titular": "contacto_titular",
        # Estado
        "estado_academico": "estado_academico",
        "estado": "estado_academico",
        "estado_est": "estado_academico",
    }

    name = fields.Char(
        string="Referencia",
        default=lambda self: self.env["ir.sequence"].next_by_code(
            "benglish.student.import.batch"
        )
        or _("Importación"),
        tracking=True,
    )
    state = fields.Selection(
        [
            ("draft", "Borrador"),
            ("validated", "Validado"),
            ("ready", "Listo"),
            ("imported", "Importado"),
            ("error", "Error"),
        ],
        string="Estado",
        default="draft",
        tracking=True,
        required=True,
    )
    file_name = fields.Char(string="Archivo")
    file_data = fields.Binary(string="Archivo XLSX", attachment=True)
    file_checksum = fields.Char(string="Hash", readonly=True)
    file_size = fields.Integer(string="Tamaño", readonly=True)
    file_mimetype = fields.Char(string="MIME", readonly=True)

    header_snapshot = fields.Json(string="Snapshot columnas")
    extra_columns = fields.Text(string="Columnas extra")
    missing_columns = fields.Text(string="Columnas faltantes")

    extra_column_policy = fields.Selection(
        [
            ("reject", "Rechazar"),
            ("ignore", "Ignorar"),
            ("allow", "Permitir"),
        ],
        string="Política de columnas extra",
        default="reject",
        required=True,
    )
    enforce_column_order = fields.Boolean(
        string="Validar el orden de columnas", default=False
    )
    identity_field = fields.Selection(
        [("student_id_number", "Documento de identidad")],
        string="Campo de identidad",
        default="student_id_number",
        required=True,
    )

    line_ids = fields.One2many(
        "benglish.student.import.line",
        "batch_id",
        string="Líneas",
    )
    log_ids = fields.One2many("benglish.student.import.log", "batch_id", string="Logs")

    validated_by = fields.Many2one("res.users", string="Validado por", readonly=True)
    validated_at = fields.Datetime(string="Validado el", readonly=True)
    ready_by = fields.Many2one("res.users", string="Listo por", readonly=True)
    ready_at = fields.Datetime(string="Listo el", readonly=True)
    imported_by = fields.Many2one("res.users", string="Importado por", readonly=True)
    imported_at = fields.Datetime(string="Importado el", readonly=True)

    line_count = fields.Integer(compute="_compute_counts", string="Total de líneas")
    error_count = fields.Integer(compute="_compute_counts", string="Errores")
    warning_count = fields.Integer(compute="_compute_counts", string="Advertencias")
    pending_decision_count = fields.Integer(
        compute="_compute_counts", string="Pendientes"
    )
    imported_count = fields.Integer(compute="_compute_counts", string="Procesadas")

    error_report_file = fields.Binary(
        string="Reporte de errores", attachment=True, readonly=True
    )
    error_report_filename = fields.Char(string="Nombre reporte", readonly=True)

    @api.depends(
        "line_ids",
        "line_ids.validation_state",
        "line_ids.action_decision",
        "line_ids.result_state",
    )
    def _compute_counts(self):
        for batch in self:
            lines = batch.line_ids
            batch.line_count = len(lines)
            batch.error_count = len(
                lines.filtered(lambda l: l.validation_state == "error")
            )
            batch.warning_count = len(
                lines.filtered(lambda l: l.validation_state == "warning")
            )
            batch.pending_decision_count = len(
                lines.filtered(lambda l: l.action_decision == "pending")
            )
            batch.imported_count = len(
                lines.filtered(
                    lambda l: l.result_state in ("created", "updated", "ignored")
                )
            )

    def _log(self, level, message, line=None, details=None):
        self.ensure_one()
        return self.env["benglish.student.import.log"].create(
            {
                "batch_id": self.id,
                "line_id": line.id if line else False,
                "level": level,
                "message": message,
                "details": details,
            }
        )

    def _ensure_xlsx(self):
        self.ensure_one()
        if not self.file_data or not self.file_name:
            raise UserError(_("Debe cargar un archivo XLSX antes de continuar."))
        if not self.file_name.lower().endswith(".xlsx"):
            raise UserError(_("Formato inválido. Solo se permiten archivos .xlsx."))

    def _load_workbook(self):
        if openpyxl is None:
            raise UserError(
                _(
                    "La libreria openpyxl no esta disponible. "
                    "Instalela para procesar archivos XLSX."
                )
            )
        self._ensure_xlsx()
        file_bytes = base64.b64decode(self.file_data)
        return openpyxl.load_workbook(
            io.BytesIO(file_bytes), data_only=True, read_only=True
        )

    @api.model
    def _normalize_header(self, value):
        if value is None:
            return ""
        text = str(value).strip()
        if not text:
            return ""
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = text.lower()
        text = re.sub(r"[\s\-\/]+", "_", text)
        text = re.sub(r"[^a-z0-9_]", "", text)
        return text

    @api.model
    def _normalize_key(self, value):
        text = self._cell_to_string(value)
        if not text:
            return ""
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        text = re.sub(r"\s+", " ", text.strip().lower())
        return text

    @api.model
    def _cell_to_string(self, value):
        if value is None:
            return False
        if isinstance(value, bool):
            return "1" if value else "0"
        if isinstance(value, (int,)):
            return str(value)
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value)
        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d")
        text = str(value).strip()
        return text or False

    @api.model
    def _normalize_email(self, value):
        text = self._cell_to_string(value)
        if not text:
            return False
        normalized = email_normalize(text)
        return normalized or text.strip().lower()

    @api.model
    def _normalize_phone(self, value):
        text = self._cell_to_string(value)
        if not text:
            return False
        text = re.sub(r"[^\d+]", "", text)
        return text or False

    @api.model
    def _parse_date(self, value):
        if not value:
            return (False, False)
        if isinstance(value, datetime):
            return (value.date(), False)
        if isinstance(value, date):
            return (value, False)
        text = self._cell_to_string(value)
        if not text:
            return (False, False)
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return (datetime.strptime(text, fmt).date(), False)
            except ValueError:
                continue
        return (False, True)

    @api.model
    def _normalize_gender(self, value):
        token = self._normalize_header(value)
        if not token:
            return (False, False)
        mapping = {
            "m": "male",
            "masculino": "male",
            "male": "male",
            "f": "female",
            "femenino": "female",
            "female": "female",
            "otro": "other",
            "other": "other",
            "o": "other",
        }
        if token in mapping:
            return (mapping[token], False)
        return (False, True)

    @api.model
    def _normalize_delivery_mode(self, value):
        token = self._normalize_header(value)
        if not token:
            return (False, False)
        mapping = {
            "presencial": "presential",
            "presential": "presential",
            "virtual": "virtual",
            "hibrido": "hybrid",
            "hibrido_presencial": "hybrid",
            "hybrid": "hybrid",
        }
        if token in mapping:
            return (mapping[token], False)
        return (False, True)

    @api.model
    def _normalize_estado_academico(self, value):
        """Normaliza el estado académico del estudiante"""
        token = self._normalize_header(value)
        if not token:
            return (False, False)
        mapping = {
            "prospecto": "prospect",
            "prospect": "prospect",
            "matriculado": "enrolled",
            "enrolled": "enrolled",
            "activo": "active",
            "active": "active",
            "inactivo": "inactive",
            "inactive": "inactive",
            "graduado": "graduated",
            "graduated": "graduated",
            "retirado": "withdrawn",
            "withdrawn": "withdrawn",
        }
        if token in mapping:
            return (mapping[token], False)
        return (False, True)

    @api.model
    def _json_safe(self, value):
        if value is None:
            return False
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, bool):
            return value
        return str(value)

    def _validate_headers(self, header_row):
        expected = self.REQUIRED_COLUMNS + self.OPTIONAL_COLUMNS
        header_snapshot = []
        canonical_headers = []
        extra_headers = []
        duplicates = []
        seen = set()

        for original in header_row:
            normalized = self._normalize_header(original)
            canonical = self.COLUMN_ALIASES.get(normalized)
            header_snapshot.append(
                {
                    "original": original,
                    "normalized": normalized,
                    "canonical": canonical,
                }
            )
            if not canonical:
                if normalized:
                    extra_headers.append(original)
                continue
            if canonical in seen:
                duplicates.append(canonical)
            else:
                seen.add(canonical)
                canonical_headers.append(canonical)

        missing = [col for col in self.REQUIRED_COLUMNS if col not in seen]

        errors = []
        if missing:
            errors.append(_("Columnas faltantes: %s") % ", ".join(missing))
        if duplicates:
            errors.append(
                _("Columnas duplicadas detectadas: %s") % ", ".join(duplicates)
            )
        if extra_headers and self.extra_column_policy == "reject":
            errors.append(
                _("Columnas extra no permitidas: %s")
                % ", ".join([str(col) for col in extra_headers])
            )

        if self.enforce_column_order:
            ordered = [col for col in expected if col in canonical_headers]
            if canonical_headers != ordered:
                errors.append(
                    _(
                        "El orden de columnas no coincide con el esperado. "
                        "Revise la plantilla oficial."
                    )
                )

        return {
            "errors": errors,
            "missing": missing,
            "extra": extra_headers,
            "snapshot": header_snapshot,
            "canonical_headers": canonical_headers,
        }

    @api.model
    def _normalize_for_match(self, value):
        """
        Normaliza un valor para comparación case-insensitive.
        Elimina acentos, espacios extra, y convierte a minúsculas.
        """
        if not value:
            return ""
        text = str(value).strip()
        # Eliminar acentos
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        # Convertir a minúsculas y normalizar espacios
        text = re.sub(r"\s+", " ", text.lower())
        return text

    def _match_by_name(self, model_name, value, cache):
        """
        Busca un registro por nombre de forma case-insensitive y flexible.
        Soporta variaciones como 'BENGLISH', 'benglish', 'Benglish', etc.
        """
        if not value:
            return (False, False)
        
        # Normalizar el valor de búsqueda
        normalized_search = self._normalize_for_match(value)
        if not normalized_search:
            return (False, False)
        
        # Verificar en cache
        cache_key = f"{model_name}:{normalized_search}"
        if cache_key in cache:
            return cache[cache_key]

        # Buscar todos los registros activos y comparar con normalización
        all_records = self.env[model_name].search([("active", "=", True)])
        matching_records = []
        
        for record in all_records:
            record_normalized = self._normalize_for_match(record.name)
            # Comparación exacta (case-insensitive, sin acentos)
            if record_normalized == normalized_search:
                matching_records.append(record)
                continue
            # Comparación parcial (contenido)
            if normalized_search in record_normalized or record_normalized in normalized_search:
                matching_records.append(record)
        
        if not matching_records:
            # Intentar con búsqueda ilike como fallback
            records = self.env[model_name].search([("name", "ilike", value)], limit=2)
            if not records:
                cache[cache_key] = (False, "not_found")
            elif len(records) > 1:
                cache[cache_key] = (False, "ambiguous")
            else:
                cache[cache_key] = (records[0], False)
        elif len(matching_records) > 1:
            # Si hay múltiples coincidencias exactas, preferir la exacta
            exact_matches = [r for r in matching_records 
                           if self._normalize_for_match(r.name) == normalized_search]
            if len(exact_matches) == 1:
                cache[cache_key] = (exact_matches[0], False)
            else:
                cache[cache_key] = (False, "ambiguous")
        else:
            cache[cache_key] = (matching_records[0], False)
        
        return cache[cache_key]

    def _match_commercial_plan(self, value, program_id, cache):
        """
        Busca un plan comercial por nombre de forma case-insensitive.
        Soporta variaciones como 'Plan GOLD', 'plan gold', 'PLAN Gold', etc.
        Si se proporciona program_id, filtra por ese programa.
        """
        if not value:
            return (False, False)
        
        normalized_search = self._normalize_for_match(value)
        if not normalized_search:
            return (False, False)
        
        # Cache key incluye program_id
        cache_key = f"commercial_plan:{normalized_search}:{program_id or 'all'}"
        if cache_key in cache:
            return cache[cache_key]

        # Construir dominio de búsqueda
        domain = [("state", "=", "active")]
        if program_id:
            domain.append(("program_id", "=", program_id))

        all_plans = self.env["benglish.commercial.plan"].search(domain)
        matching_plans = []
        
        for plan in all_plans:
            plan_normalized = self._normalize_for_match(plan.name)
            # Comparación exacta
            if plan_normalized == normalized_search:
                matching_plans.append(plan)
                continue
            # Comparación parcial - el nombre buscado contiene el nombre del plan
            # o viceversa (ej: "plan gold" vs "gold")
            if normalized_search in plan_normalized or plan_normalized in normalized_search:
                matching_plans.append(plan)
        
        if not matching_plans:
            cache[cache_key] = (False, "not_found")
        elif len(matching_plans) > 1:
            # Preferir coincidencia exacta
            exact = [p for p in matching_plans 
                    if self._normalize_for_match(p.name) == normalized_search]
            if len(exact) == 1:
                cache[cache_key] = (exact[0], False)
            else:
                cache[cache_key] = (False, "ambiguous")
        else:
            cache[cache_key] = (matching_plans[0], False)
        
        return cache[cache_key]

    def _parse_starting_level(self, value):
        """
        Parsea el nivel de inicio desde el Excel.
        Acepta números enteros (1-24), nombres de fase (INTERMEDIATE), 
        o combinaciones (INTERMEDIATE 11 → nivel 11)
        """
        if not value:
            return (None, False)
        
        text = str(value).strip()
        
        # Intentar parsear como número directamente
        try:
            level = int(float(text))
            if 1 <= level <= 24:
                return (level, False)
            else:
                return (None, True)  # Fuera de rango
        except (ValueError, TypeError):
            pass
        
        # Intentar extraer número de una cadena (ej: "INTERMEDIATE 11" → 11)
        match = re.search(r"(\d+)", text)
        if match:
            try:
                level = int(match.group(1))
                if 1 <= level <= 24:
                    return (level, False)
            except ValueError:
                pass
        
        # No se pudo parsear
        return (None, True)

    def _match_country(self, value, cache):
        if not value:
            return (False, False)
        key = self._normalize_key(value)
        if not key:
            return (False, False)
        if key in cache:
            return cache[key]

        Country = self.env["res.country"]
        record = Country.search([("code", "=", key.upper())], limit=1)
        if not record:
            records = Country.search([("name", "ilike", value)], limit=2)
            if not records:
                cache[key] = (False, "not_found")
            elif len(records) > 1:
                cache[key] = (False, "ambiguous")
            else:
                cache[key] = (records[0], False)
        else:
            cache[key] = (record, False)
        return cache[key]

    def _build_line_vals(self, row_number, headers, row_values, caches):
        raw_values = {}
        for idx, header in enumerate(headers):
            value = row_values[idx] if idx < len(row_values) else False
            raw_values[header] = self._json_safe(value)

        is_empty = True
        for value in row_values:
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            is_empty = False
            break

        data = {}
        for idx, header in enumerate(headers):
            normalized = self._normalize_header(header)
            canonical = self.COLUMN_ALIASES.get(normalized)
            if not canonical:
                continue
            data[canonical] = row_values[idx] if idx < len(row_values) else False

        student_id_number = self._cell_to_string(data.get("documento_identidad"))
        if student_id_number:
            student_id_number = re.sub(r"\s+", "", student_id_number)

        # Nombres desagregados
        first_name = self._cell_to_string(data.get("primer_nombre"))
        second_name = self._cell_to_string(data.get("segundo_nombre"))
        first_last_name = self._cell_to_string(data.get("primer_apellido"))
        second_last_name = self._cell_to_string(data.get("segundo_apellido"))

        # Nombre completo (construido o del Excel)
        name = self._cell_to_string(data.get("nombre_completo"))
        if not name and (first_name or first_last_name):
            parts = [first_name, second_name, first_last_name, second_last_name]
            name = " ".join(part for part in parts if part).strip()

        email = self._normalize_email(data.get("email"))
        phone = self._normalize_phone(data.get("telefono"))
        mobile = self._normalize_phone(data.get("celular"))

        birth_date, birth_date_error = self._parse_date(data.get("fecha_nacimiento"))
        gender, gender_error = self._normalize_gender(data.get("genero"))
        delivery_mode, delivery_error = self._normalize_delivery_mode(
            data.get("modalidad")
        )

        code = self._cell_to_string(data.get("codigo"))
        address = self._cell_to_string(data.get("direccion"))
        city = self._cell_to_string(data.get("ciudad"))

        country_raw = self._cell_to_string(data.get("pais"))
        country, country_error = self._match_country(country_raw, caches["country"])

        program_raw = self._cell_to_string(data.get("programa"))
        program, program_error = self._match_by_name(
            "benglish.program", program_raw, caches["program"]
        )

        # Plan legacy (mantener compatibilidad)
        plan_raw = self._cell_to_string(data.get("plan"))
        plan, plan_error = self._match_by_name(
            "benglish.plan", plan_raw, caches["plan"]
        )

        # Plan comercial (Feb 2026 - nuevo modelo de matrícula)
        commercial_plan_raw = self._cell_to_string(data.get("plan_comercial"))
        commercial_plan, commercial_plan_error = self._match_commercial_plan(
            commercial_plan_raw,
            program.id if program else None,
            caches.setdefault("commercial_plan", {})
        )

        # Nivel de inicio (entero 1-24)
        level_raw = self._cell_to_string(data.get("nivel"))
        starting_level, starting_level_error = self._parse_starting_level(level_raw)

        phase_raw = self._cell_to_string(data.get("fase"))
        phase, phase_error = self._match_by_name(
            "benglish.phase", phase_raw, caches.setdefault("phase", {})
        )

        # Level como registro (para compatibilidad legacy)
        level, level_error = self._match_by_name(
            "benglish.level", level_raw, caches.setdefault("level", {})
        )

        campus_raw = self._cell_to_string(data.get("sede"))
        campus, campus_error = self._match_by_name(
            "benglish.campus", campus_raw, caches["campus"]
        )

        # Datos del contrato académico (matrícula)
        categoria = self._cell_to_string(data.get("categoria"))
        course_start_date, course_start_error = self._parse_date(
            data.get("fecha_inicio_curso")
        )
        course_end_date, course_end_error = self._parse_date(
            data.get("fecha_fin_curso")
        )
        max_freeze_date, max_freeze_error = self._parse_date(
            data.get("fecha_maxima_congelamiento")
        )

        course_days_raw = data.get("dias_curso")
        course_days = None
        if course_days_raw:
            try:
                course_days = int(float(course_days_raw)) if course_days_raw else None
            except (ValueError, TypeError):
                course_days = None

        # Estado académico
        estado_raw = self._cell_to_string(data.get("estado_academico"))
        estado, estado_error = self._normalize_estado_academico(estado_raw)

        # Titular
        responsible_name = self._cell_to_string(data.get("contacto_titular"))

        return {
            "batch_id": self.id,
            "row_number": row_number,
            "source": "xlsx",
            "raw_values": raw_values,
            "first_name": first_name,
            "second_name": second_name,
            "first_last_name": first_last_name,
            "second_last_name": second_last_name,
            "name": name,
            "student_id_number": student_id_number,
            "email": email,
            "phone": phone,
            "mobile": mobile,
            "birth_date": birth_date,
            "birth_date_parse_error": birth_date_error,
            "gender": gender,
            "gender_parse_error": gender_error,
            "code": code,
            "address": address,
            "city": city,
            "country_id": country.id if country else False,
            "country_match_error": bool(country_error),
            "program_id": program.id if program else False,
            "program_match_error": bool(program_error),
            "plan_id": plan.id if plan else False,
            "plan_match_error": bool(plan_error),
            "commercial_plan_id": commercial_plan.id if commercial_plan else False,
            "commercial_plan_match_error": bool(commercial_plan_error),
            "starting_level": starting_level,
            "starting_level_parse_error": starting_level_error,
            "phase_id": phase.id if phase else False,
            "phase_match_error": bool(phase_error),
            "level_id": level.id if level else False,
            "level_match_error": bool(level_error),
            "preferred_campus_id": campus.id if campus else False,
            "campus_match_error": bool(campus_error),
            "preferred_delivery_mode": delivery_mode,
            "delivery_mode_parse_error": delivery_error,
            "categoria": categoria,
            "course_start_date": course_start_date,
            "course_start_date_parse_error": course_start_error,
            "course_end_date": course_end_date,
            "course_end_date_parse_error": course_end_error,
            "max_freeze_date": max_freeze_date,
            "max_freeze_date_parse_error": max_freeze_error,
            "course_days": course_days,
            "estado_academico": estado,
            "estado_parse_error": estado_error,
            "responsible_name": responsible_name,
            "action_decision": "pending",
            "is_empty": is_empty,
        }

    def _prepare_file_metadata(self, file_data, file_name=None):
        file_bytes = base64.b64decode(file_data)
        values = {
            "file_size": len(file_bytes),
            "file_checksum": hashlib.sha256(file_bytes).hexdigest(),
        }
        if file_name:
            mimetype = mimetypes.guess_type(file_name)[0]
            if mimetype:
                values["file_mimetype"] = mimetype
        return values

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            file_data = vals.get("file_data")
            if file_data:
                vals.update(
                    self._prepare_file_metadata(file_data, vals.get("file_name"))
                )
        return super().create(vals_list)

    def write(self, vals):
        if "file_data" in vals:
            if any(batch.state != "draft" for batch in self):
                raise UserError(
                    _("No puede modificar el archivo fuera del estado borrador.")
                )
            file_name = vals.get("file_name")
            vals.update(self._prepare_file_metadata(vals.get("file_data"), file_name))
        return super().write(vals)

    def action_validate_file(self):
        self.ensure_one()
        if self.state != "draft":
            raise UserError(_("Solo puede validar batches en borrador."))

        workbook = self._load_workbook()
        sheet = workbook.active

        try:
            header_row = [
                cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
            ]
        except StopIteration:
            raise UserError(_("El archivo no contiene filas.")) from None
        validation = self._validate_headers(header_row)
        self.header_snapshot = validation["snapshot"]
        self.missing_columns = (
            ", ".join(validation["missing"]) if validation["missing"] else False
        )
        self.extra_columns = (
            ", ".join([str(col) for col in validation["extra"]])
            if validation["extra"]
            else False
        )

        if validation["errors"]:
            message = "\n".join("- %s" % err for err in validation["errors"])
            raise UserError(_("Estructura de archivo inválida:\n%s") % message)

        if self.line_ids:
            self.line_ids.unlink()

        caches = {
            "program": {},
            "plan": {},
            "campus": {},
            "country": {},
        }

        line_vals = []
        row_number = 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_number += 1
            line_vals.append(self._build_line_vals(row_number, header_row, row, caches))
            if len(line_vals) >= 500:
                self.env["benglish.student.import.line"].with_context(
                    skip_import_recompute=True
                ).create(line_vals)
                line_vals = []

        if line_vals:
            self.env["benglish.student.import.line"].with_context(
                skip_import_recompute=True
            ).create(line_vals)

        self._recompute_duplicates_for_keys(set(self.line_ids.mapped("identity_key")))
        self._recompute_decision_conflicts(set(self.line_ids.mapped("identity_key")))
        self._suggest_actions()

        if validation["extra"] and self.extra_column_policy == "ignore":
            self._log(
                "warning",
                _("Columnas extra ignoradas"),
                details=", ".join([str(col) for col in validation["extra"]]),
            )

        self.state = "validated"
        self.validated_by = self.env.user
        self.validated_at = fields.Datetime.now()

        self._log("info", _("Batch validado y cargado."))

    def _suggest_actions(self):
        for line in self.line_ids.filtered(
            lambda l: l.action_decision == "pending"
            and l.duplicate_scope == "none"
            and l.validation_state == "ok"
        ):
            line.with_context(skip_import_recompute=True).write(
                {"action_decision": "create"}
            )

    def _recompute_duplicates_for_keys(self, keys):
        self.ensure_one()
        keys = [key for key in keys if key]
        if not keys:
            return

        lines = self.env["benglish.student.import.line"].search(
            [("batch_id", "=", self.id), ("identity_key", "in", keys)]
        )
        if not lines:
            return

        lines.with_context(skip_import_recompute=True).write(
            {"duplicate_in_batch": False, "duplicate_line_numbers": False}
        )

        grouped = {}
        for line in lines:
            if line.identity_key:
                grouped.setdefault(line.identity_key, []).append(line)

        for group_lines in grouped.values():
            if len(group_lines) < 2:
                continue
            for line in group_lines:
                others = [
                    str(other.row_number)
                    for other in group_lines
                    if other.id != line.id and other.row_number
                ]
                line.with_context(skip_import_recompute=True).write(
                    {
                        "duplicate_in_batch": True,
                        "duplicate_line_numbers": (
                            ", ".join(others) if others else False
                        ),
                    }
                )

        students = self.env["benglish.student"].search(
            [("student_id_number", "in", keys)]
        )
        student_map = {stud.student_id_number: stud.id for stud in students}

        for line in lines:
            student_id = (
                student_map.get(line.student_id_number)
                if line.student_id_number
                else False
            )
            values = {"existing_student_id": student_id or False}
            if (
                student_id
                and line.action_decision == "update"
                and not line.target_student_id
            ):
                values["target_student_id"] = student_id
            line.with_context(skip_import_recompute=True).write(values)

    def _recompute_decision_conflicts(self, keys):
        self.ensure_one()
        keys = [key for key in keys if key]
        if not keys:
            return
        lines = self.env["benglish.student.import.line"].search(
            [("batch_id", "=", self.id), ("identity_key", "in", keys)]
        )
        if not lines:
            return

        lines.with_context(skip_import_recompute=True).write(
            {"decision_conflict": False, "decision_conflict_message": False}
        )

        grouped = {}
        for line in lines:
            if line.identity_key:
                grouped.setdefault(line.identity_key, []).append(line)

        for key, group_lines in grouped.items():
            if len(group_lines) < 2:
                continue
            chosen = [
                line
                for line in group_lines
                if line.action_decision in ("create", "update")
            ]
            if len(chosen) > 1:
                message = _(
                    "Hay multiples acciones de crear/actualizar para la misma identidad."
                )
                for line in group_lines:
                    line.with_context(skip_import_recompute=True).write(
                        {
                            "decision_conflict": True,
                            "decision_conflict_message": message,
                        }
                    )

    def action_set_ready(self):
        self.ensure_one()
        if self.state != "validated":
            raise UserError(_("Solo puede marcar listo desde estado validado."))

        if self.error_count:
            raise UserError(
                _("No puede continuar. Existen errores en las líneas de staging.")
            )
        if self.pending_decision_count:
            raise UserError(
                _("No puede continuar. Hay decisiones pendientes en líneas duplicadas.")
            )
        if self.line_ids.filtered(lambda l: l.decision_conflict):
            raise UserError(
                _(
                    "No puede continuar. Hay conflictos de decisión en líneas duplicadas."
                )
            )

        self.state = "ready"
        self.ready_by = self.env.user
        self.ready_at = fields.Datetime.now()
        self._log("info", _("Batch marcado como listo para importar."))

    def action_open_confirm_wizard(self):
        self.ensure_one()
        return {
            "name": _("Confirmar importación"),
            "type": "ir.actions.act_window",
            "res_model": "benglish.student.import.confirm.wizard",
            "view_mode": "form",
            "target": "new",
            "context": {"default_batch_id": self.id},
        }

    def action_open_bulk_decision_wizard(self):
        self.ensure_one()
        if self.state not in ("validated", "ready"):
            raise UserError(
                _("Solo puede cambiar decisiones en estado validado o listo.")
            )
        return {
            "name": _("Cambiar decisiones"),
            "type": "ir.actions.act_window",
            "res_model": "benglish.student.import.bulk.decision.wizard",
            "view_mode": "form",
            "target": "new",
            "context": {"default_batch_id": self.id},
        }

    def _prepare_student_vals(self, line):
        """Prepara los valores para crear/actualizar un estudiante."""
        # Nombres desagregados - el campo 'name' es computado, no se pasa directamente
        vals = {
            "first_name": line.first_name,
            "second_name": line.second_name,
            "first_last_name": line.first_last_name,
            "second_last_name": line.second_last_name,
            # "name": se computa automáticamente desde los campos anteriores
            "student_id_number": line.student_id_number,
            "email": line.email,
            "phone": line.phone,
            "mobile": line.mobile,
            "birth_date": line.birth_date,
            "gender": line.gender,
            "code": line.code,
            "address": line.address,
            "city": line.city,
            "country_id": line.country_id.id if line.country_id else False,
            "program_id": line.program_id.id if line.program_id else False,
            "plan_id": line.plan_id.id if line.plan_id else False,
            "commercial_plan_id": (
                line.commercial_plan_id.id if line.commercial_plan_id else False
            ),
            "preferred_campus_id": (
                line.preferred_campus_id.id if line.preferred_campus_id else False
            ),
            "preferred_delivery_mode": line.preferred_delivery_mode,
        }
        return {
            key: value for key, value in vals.items() if value not in (False, None, "")
        }

    def _prepare_enrollment_vals(self, student, line):
        """
        Prepara los valores para crear una matrícula automáticamente.
        
        Feb 2026 - Modelo de matrícula por nivel:
        - program_id: obligatorio
        - commercial_plan_id: obligatorio (define estructura de asignaturas)
        - current_level: nivel numérico de inicio (1-24)
        """
        vals = {
            "student_id": student.id,
            "program_id": line.program_id.id if line.program_id else False,
            "commercial_plan_id": (
                line.commercial_plan_id.id if line.commercial_plan_id else False
            ),
            "delivery_mode": line.preferred_delivery_mode or "presential",
            "state": "active",
        }
        
        # Nivel de inicio
        if line.starting_level:
            vals["current_level"] = line.starting_level
            vals["starting_level"] = line.starting_level
        elif line.commercial_plan_id:
            # Usar el nivel inicial del plan comercial
            vals["current_level"] = line.commercial_plan_id.level_start or 1
            vals["starting_level"] = line.commercial_plan_id.level_start or 1
        
        # Datos adicionales del contrato
        if line.categoria:
            vals["categoria"] = line.categoria
        if line.course_start_date:
            vals["course_start_date"] = line.course_start_date
        if line.course_end_date:
            vals["course_end_date"] = line.course_end_date
        if line.max_freeze_date:
            vals["max_freeze_date"] = line.max_freeze_date
        if line.course_days:
            vals["course_days"] = line.course_days
        
        # Plan legacy (compatibilidad)
        if line.plan_id:
            vals["plan_id"] = line.plan_id.id
        
        # Fase y nivel como registros (compatibilidad)
        if line.phase_id:
            vals["current_phase_id"] = line.phase_id.id
        if line.level_id:
            vals["current_level_id"] = line.level_id.id
            
        return vals

    def action_import(self):
        self.ensure_one()
        if self.state != "ready":
            raise UserError(_("Solo puede importar batches en estado listo."))

        lines = self.line_ids.filtered(
            lambda l: l.action_decision in ("create", "update", "ignore")
            and l.result_state in ("pending", "error")
        )
        if not lines:
            raise UserError(_("No hay líneas pendientes para procesar."))

        Student = self.env["benglish.student"]
        errors_found = False

        for line in lines:
            if line.validation_state == "error":
                errors_found = True
                line.with_context(skip_import_recompute=True).write(
                    {
                        "result_state": "error",
                        "result_message": _("Línea con errores de validación."),
                    }
                )
                continue

            try:
                with self.env.cr.savepoint():
                    if line.action_decision == "ignore":
                        line.with_context(skip_import_recompute=True).write(
                            {
                                "result_state": "ignored",
                                "result_message": _(
                                    "Ignorado por decisión del usuario."
                                ),
                                "processed_at": fields.Datetime.now(),
                                "processed_by": self.env.user.id,
                            }
                        )
                        self._log("info", _("Línea ignorada."), line=line)
                        continue

                    vals = self._prepare_student_vals(line)

                    if line.action_decision == "create":
                        student = Student.create(vals)
                        
                        # ═══════════════════════════════════════════════════════════
                        # CREACIÓN AUTOMÁTICA DE MATRÍCULA (Feb 2026)
                        # ═══════════════════════════════════════════════════════════
                        enrollment = None
                        enrollment_message = ""
                        
                        if line.program_id and line.commercial_plan_id:
                            try:
                                enrollment_vals = self._prepare_enrollment_vals(student, line)
                                enrollment = self.env["benglish.enrollment"].create(enrollment_vals)
                                enrollment_message = _(
                                    " Matrícula creada: %s (Nivel %d)"
                                ) % (enrollment.code, enrollment.current_level)
                                _logger.info(
                                    f"[IMPORT] Matrícula creada automáticamente: "
                                    f"estudiante={student.name}, "
                                    f"programa={line.program_id.name}, "
                                    f"plan={line.commercial_plan_id.name}, "
                                    f"nivel={enrollment.current_level}"
                                )
                            except Exception as enroll_exc:
                                enrollment_message = _(
                                    " Error al crear matrícula: %s"
                                ) % str(enroll_exc)
                                _logger.warning(
                                    f"[IMPORT] Error al crear matrícula para {student.name}: {enroll_exc}"
                                )
                        elif line.program_id or line.commercial_plan_id:
                            enrollment_message = _(
                                " Matrícula no creada: faltan datos (programa y/o plan comercial)."
                            )
                        
                        result_message = _("Estudiante creado.") + enrollment_message
                        
                        line.with_context(skip_import_recompute=True).write(
                            {
                                "result_state": "created",
                                "result_message": result_message,
                                "processed_at": fields.Datetime.now(),
                                "processed_by": self.env.user.id,
                            }
                        )
                        self._log("info", result_message, line=line)
                        line.with_context(skip_import_recompute=True).write(
                            {"existing_student_id": student.id}
                        )
                        continue

                    if line.action_decision == "update":
                        target = line.target_student_id
                        if not target:
                            raise UserError(
                                _("No hay estudiante objetivo para actualizar.")
                            )
                        target.write(vals)
                        
                        # ═══════════════════════════════════════════════════════════
                        # ACTUALIZACIÓN/CREACIÓN DE MATRÍCULA AL ACTUALIZAR ESTUDIANTE
                        # ═══════════════════════════════════════════════════════════
                        enrollment_message = ""
                        
                        if line.program_id and line.commercial_plan_id:
                            # Verificar si ya tiene matrícula activa
                            existing_enrollment = self.env["benglish.enrollment"].search([
                                ("student_id", "=", target.id),
                                ("program_id", "=", line.program_id.id),
                                ("commercial_plan_id", "=", line.commercial_plan_id.id),
                                ("state", "in", ["active", "enrolled", "in_progress"]),
                            ], limit=1)
                            
                            if existing_enrollment:
                                # Actualizar nivel si se especifica uno diferente
                                if line.starting_level and existing_enrollment.current_level != line.starting_level:
                                    existing_enrollment.write({"current_level": line.starting_level})
                                    enrollment_message = _(
                                        " Matrícula existente actualizada: %s (Nivel %d)"
                                    ) % (existing_enrollment.code, line.starting_level)
                                else:
                                    enrollment_message = _(
                                        " Matrícula existente: %s"
                                    ) % existing_enrollment.code
                            else:
                                # Crear nueva matrícula
                                try:
                                    enrollment_vals = self._prepare_enrollment_vals(target, line)
                                    enrollment = self.env["benglish.enrollment"].create(enrollment_vals)
                                    enrollment_message = _(
                                        " Nueva matrícula creada: %s (Nivel %d)"
                                    ) % (enrollment.code, enrollment.current_level)
                                except Exception as enroll_exc:
                                    enrollment_message = _(
                                        " Error al crear matrícula: %s"
                                    ) % str(enroll_exc)
                        
                        result_message = _("Estudiante actualizado.") + enrollment_message
                        
                        line.with_context(skip_import_recompute=True).write(
                            {
                                "result_state": "updated",
                                "result_message": result_message,
                                "processed_at": fields.Datetime.now(),
                                "processed_by": self.env.user.id,
                            }
                        )
                        self._log("info", result_message, line=line)

            except Exception as exc:  # pragma: no cover - audit trail
                errors_found = True
                _logger.exception("Error procesando linea %s", line.row_number)
                line.with_context(skip_import_recompute=True).write(
                    {
                        "result_state": "error",
                        "result_message": str(exc),
                        "processed_at": fields.Datetime.now(),
                        "processed_by": self.env.user.id,
                    }
                )
                self._log("error", _("Error en línea."), line=line, details=str(exc))

        self.imported_by = self.env.user
        self.imported_at = fields.Datetime.now()
        self.state = "error" if errors_found else "imported"
        self._log(
            "warning" if errors_found else "info",
            (
                _("Importación finalizada con errores.")
                if errors_found
                else _("Importación completada exitosamente.")
            ),
        )

    def action_generate_error_report(self):
        self.ensure_one()
        if openpyxl is None:
            raise UserError(
                _(
                    "La libreria openpyxl no esta disponible. "
                    "Instalela para generar reportes XLSX."
                )
            )

        lines = self.line_ids.filtered(
            lambda l: l.validation_state == "error" or l.result_state == "error"
        )
        if not lines:
            raise UserError(_("No hay errores para reportar."))

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Errores"

        headers = [
            "Fila",
            "Documento",
            "Nombre",
            "Correo",
            "Teléfono",
            "Decisión",
            "Estado de validación",
            "Errores",
            "Resultado",
            "Detalle del resultado",
        ]
        sheet.append(headers)

        for line in lines:
            sheet.append(
                [
                    line.row_number,
                    line.student_id_number,
                    line.name,
                    line.email,
                    line.phone,
                    line.action_decision,
                    line.validation_state,
                    line.error_messages,
                    line.result_state,
                    line.result_message,
                ]
            )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        content = base64.b64encode(buffer.read())

        filename = "%s-errores.xlsx" % (self.name or "importacion")
        self.write(
            {
                "error_report_file": content,
                "error_report_filename": filename,
            }
        )

        self._log("info", _("Reporte de errores generado."))

    def action_open_form(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "res_model": "benglish.student.import.batch",
            "view_mode": "form",
            "res_id": self.id,
            "target": "current",
        }
