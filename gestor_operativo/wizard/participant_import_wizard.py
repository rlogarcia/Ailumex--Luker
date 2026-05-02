# -*- coding: utf-8 -*-
# VERSION: 2026-04-06-v3 — OPE_Carga_Poblacion field names
import base64
import io
import json
import logging
from datetime import datetime

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────
# CAMPOS DESTINO — con descripción de campo de contacto destino
# ─────────────────────────────────────────────────────────────────
DESTINATION_FIELDS = [
    ('_skip', '— Omitir esta columna —'),

    # ── Identificación ──────────────────────────────────────────
    ('doc_number',       'Número de documento  →  Contacto: Nº identificación / VAT'),
    ('doc_type_code',    'Tipo doc. código SIMAT  →  Contacto: Tipo identificación'),
    ('doc_type_name',    'Tipo doc. texto  →  Contacto: Tipo identificación (por nombre)'),

    # ── Nombre → contacto ───────────────────────────────────────
    ('first_name',       'Primer nombre  →  Contacto: Primer nombre'),
    ('second_name',      'Segundo nombre / Otros nombres  →  Contacto: Otros nombres'),
    ('first_surname',    'Primer apellido  →  Contacto: Primer apellido'),
    ('second_surname',   'Segundo apellido  →  Contacto: Segundo apellido'),

    # ── Datos personales ────────────────────────────────────────
    ('birthdate',        'Fecha de nacimiento  →  Contacto: Fecha de nacimiento'),
    ('age',              'Edad  →  Contacto: Edad (calculada, solo referencia)'),
    ('gender',           'Género (M/F)  →  Contacto: Sexo biológico / Género'),
    ('civil_state',      'Estado civil  →  Contacto: Estado civil'),

    # ── Ubicación y residencia ──────────────────────────────────
    ('address',          'Dirección de residencia  →  Contacto: Dirección'),
    ('zone',             'Zona (1=Urbana, 2=Rural)  →  Contacto: Zona'),
    ('res_depto',        'Departamento residencia (código)  →  Contacto: Departamento'),
    ('res_mun',          'Municipio residencia (código)  →  Contacto: Ciudad'),
    ('phone',            'Teléfono / Celular  →  Contacto: Teléfono'),

    # ── Nacimiento ──────────────────────────────────────────────
    ('nac_depto',        'Depto nacimiento (código)  →  Contacto: País de nacimiento (ref)'),
    ('nac_mun',          'Municipio nacimiento (código)  →  Contacto: referencia'),
    ('pais_origen',      'Código país de origen  →  Contacto: País de nacimiento'),
    ('nombre_pais',      'Nombre país de origen  →  Contacto: País de nacimiento (texto)'),
    ('exp_depto',        'Depto expedición doc.  →  Contacto: Lugar de expedición'),

    # ── Socioeconómico → contacto / atributos ───────────────────
    ('estrato',          'Estrato socioeconómico  →  Contacto: Estrato + attr_estrato'),
    ('sisben',           'Clasificación SISBEN IV  →  Contacto: SISBEN + attr_sisben'),
    ('sisben_bool',      'SISBEN (S/N)  →  Contacto: SISBEN (activar checkbox)'),

    # ── Discapacidad ────────────────────────────────────────────
    ('disability_code',  'Código discapacidad  →  Contacto: Discapacidad (activar si ≠99)'),
    ('disability_name',  'Nombre discapacidad  →  Contacto: Tipo de discapacidad'),

    # ── Etnia y conflicto ───────────────────────────────────────
    ('etnia_code',       'Código etnia  →  attr_etnia + Contacto'),
    ('etnia_name',       'Nombre etnia  →  referencia'),
    ('victim_conflict',  'Víctima conflicto (SI/NO)  →  Contacto: Víctima conflicto + attr'),

    # ── Contexto educativo ──────────────────────────────────────
    ('institution_name', 'Nombre institución  →  Organización'),
    ('sede_name',        'Nombre sede  →  Organización: sede'),
    ('institution_dane', 'Código DANE institución  →  Organización: código DANE'),
    ('grade',            'Grado  →  Unidad organizacional: grado'),
    ('group_name',       'Grupo  →  Unidad organizacional: grupo'),
    ('jornada_code',     'Código jornada  →  Unidad: jornada'),
    ('jornada_name',     'Nombre jornada  →  Unidad: jornada'),

    # ── Otros atributos ─────────────────────────────────────────
    ('repeating',        'Repitente (S/N)  →  attr_repitente'),
    ('zone_attr',        'Zona alumno  →  attr zona'),
]

# Sugerencias automáticas columna SIMAT → campo destino
SIMAT_AUTO_MAP = {
    'NRO_DOCUMENTO':        'doc_number',
    'TIPO_DOCUMENTO':       'doc_type_code',
    'NOMPRE TIPO DOC':      'doc_type_name',
    'APELLIDO1':            'first_surname',
    'APELLIDO2':            'second_surname',
    ' NOMBRE1':             'first_name',
    'NOMBRE1':              'first_name',
    'NOMBRE2':              'second_name',
    'FECHA_NACIMIENTO':     'birthdate',
    'EDAD':                 'age',
    'GENERO':               'gender',
    'TEL':                  'phone',
    'DIRECCION_RESIDENCIA': 'address',
    'ZON_ALU':              'zone',
    'RES_DEPTO':            'res_depto',
    'RES_MUN':              'res_mun',
    'NAC_DEPTO':            'nac_depto',
    'NAC_MUN':              'nac_mun',
    'PAIS_ORIGEN':          'pais_origen',
    'NOMBRE PAIS ORIGEN':   'nombre_pais',
    'EXP_DEPTO':            'exp_depto',
    'CODIGO_DANE':          'institution_dane',
    'DANE_ANTERIOR':        'institution_dane',
    'INSTITUCION':          'institution_name',
    'SEDE':                 'sede_name',
    'GRADO':                'grade',
    'GRUPO':                'group_name',
    'TIPO_JORNADA':         'jornada_code',
    'NOMBRE JORNADA':       'jornada_name',
    'ESTRATO':              'estrato',
    'SISBEN IV':            'sisben',
    'SISBEN':               'sisben',
    'TIPO_DISCAPACIDAD':    'disability_code',
    'NOMBRE DISCAPACIDAD':  'disability_name',
    'ETNIA':                'etnia_code',
    'NOMBRE ETNIA':         'etnia_name',
    'POB_VICT_CONF_RUV':    'victim_conflict',
    'REPITENTE':            'repeating',
}

IMPORTANT_COLUMNS = {'doc_number', 'first_surname', 'first_name', 'grade', 'institution_name'}

# Código SIMAT → nombre tipo identificación en l10n_latam
SIMAT_DOC_TYPE_NAMES = {
    1: 'Registro Civil',         '1': 'Registro Civil',
    2: 'Tarjeta de Identidad',   '2': 'Tarjeta de Identidad',
    3: 'Cédula de Ciudadanía',   '3': 'Cédula de Ciudadanía',
    4: 'Tarjeta de Extranjería', '4': 'Tarjeta de Extranjería',
    5: 'Cédula de Extranjería',  '5': 'Cédula de Extranjería',
    6: 'NIT',                    '6': 'NIT',
    7: 'Pasaporte',              '7': 'Pasaporte',
    11: 'NUIP',                  '11': 'NUIP',
    12: 'NIT',                   '12': 'NIT',
    13: 'Cédula de Ciudadanía',  '13': 'Cédula de Ciudadanía',
}

GENDER_MAP = {
    'M': 'masculino', 'H': 'masculino', 'MASCULINO': 'masculino', '1': 'masculino',
    'F': 'femenino',  'FEMENINO': 'femenino', '2': 'femenino',
}
JORNADA_MAP = {
    '1': 'manana', '2': 'manana', '3': 'tarde', '4': 'noche', '5': 'unica', '6': 'unica',
    1: 'manana', 2: 'manana', 3: 'tarde', 4: 'noche', 5: 'unica', 6: 'unica',
    'MAÑANA': 'manana', 'MANANA': 'manana', 'TARDE': 'tarde',
    'NOCHE': 'noche', 'ÚNICA': 'unica', 'UNICA': 'unica',
}
ZONE_MAP = {'1': 'urbana', '2': 'rural', 1: 'urbana', 2: 'rural',
            'URBANA': 'urbana', 'RURAL': 'rural'}

# ─────────────────────────────────────────────────────────────────
# Candidatos de campos en res.partner — se prueban en orden
# ─────────────────────────────────────────────────────────────────
PARTNER_FIELD_CANDIDATES = {
    'first_name':     ['first_name', 'firstname', 'l10n_latam_first_name', 'x_first_name',
                       'primer_nombre', 'x_primer_nombre'],
    'second_name':    ['other_name', 'second_name', 'middlename', 'l10n_latam_second_name',
                       'x_other_name', 'otros_nombres', 'x_otros_nombres', 'segundo_nombre'],
    'first_surname':  ['first_lastname', 'lastname', 'l10n_latam_first_lastname',
                       'x_first_lastname', 'primer_apellido', 'x_primer_apellido'],
    'second_surname': ['second_lastname', 'l10n_latam_second_lastname',
                       'x_second_lastname', 'segundo_apellido', 'x_segundo_apellido'],
    'birthdate':      ['birthdate', 'birth_date', 'x_birthdate', 'fecha_nacimiento',
                       'x_fecha_nacimiento'],
    'gender_bio':     ['biological_sex', 'x_biological_sex', 'gender', 'sex',
                       'sexo_biologico', 'x_sexo_biologico', 'x_gender'],
    'gender_id':      ['gender_id', 'x_gender_id', 'sex_id', 'identification_gender',
                       'sexo_identificacion', 'x_sexo_identificacion'],
    'civil_state':    ['marital', 'civil_state', 'estado_civil', 'x_civil_state',
                       'x_estado_civil'],
    'street':         ['street'],
    'zone':           ['zone', 'zona', 'x_zone', 'x_zona'],
    'sisben_bool':    ['sisben', 'x_sisben', 'has_sisben', 'x_has_sisben'],
    'sisben_class':   ['sisben_classification', 'x_sisben_classification', 'sisben_iv',
                       'x_sisben_iv', 'clasificacion_sisben', 'x_clasificacion_sisben'],
    'disability':     ['disability', 'x_disability', 'discapacidad', 'x_discapacidad',
                       'has_disability', 'x_has_disability'],
    'disability_type':['disability_type', 'x_disability_type', 'tipo_discapacidad',
                       'x_tipo_discapacidad'],
    'victim':         ['victim_conflict', 'x_victim_conflict', 'victima_conflicto',
                       'x_victima_conflicto', 'pob_vict_conf_ruv', 'x_pob_vict_conf_ruv'],
    'estrato':        ['stratum', 'x_stratum', 'estrato', 'x_estrato',
                       'socioeconomic_stratum', 'x_socioeconomic_stratum',
                       'estrato_socioeconomico', 'x_estrato_socioeconomico'],
    'exp_place':      ['expedition_place', 'x_expedition_place', 'lugar_expedicion',
                       'x_lugar_expedicion'],
    'id_type':        ['l10n_latam_identification_type_id'],
    'doc_number':     ['l10n_latam_document_number', 'vat'],
    'phone':          ['phone'],
    'mobile':         ['mobile'],
}

# Patrones en etiquetas de campo (field.string) para detección por label
LABEL_PATTERNS = {
    'first_name':     ['primer nombre', 'first name', 'nombre 1', 'nombre1'],
    'second_name':    ['otros nombres', 'second name', 'nombre 2', 'nombre2', 'otro nombre'],
    'first_surname':  ['primer apellido', 'first lastname', 'apellido 1', 'apellido1'],
    'second_surname': ['segundo apellido', 'second lastname', 'apellido 2', 'apellido2'],
    'estrato':        ['estrato', 'stratum', 'estrato socioeconómico'],
    'sisben_bool':    ['sisben'],
    'sisben_class':   ['clasificación sisben', 'sisben iv', 'clasificacion sisben'],
    'disability':     ['discapacidad'],
    'disability_type':['tipo de discapacidad'],
    'victim':         ['víctima', 'victima', 'conflicto'],
    'zone':           ['zona'],
    'gender_bio':     ['sexo biológico', 'sexo biologico', 'biological sex'],
}


def _resolve_partner_fields(partner_obj):
    """
    Descubre campos de res.partner en dos pasos:
    1. Por nombre técnico (PARTNER_FIELD_CANDIDATES)
    2. Por etiqueta visible (LABEL_PATTERNS) — captura campos x_studio_*
    """
    available = partner_obj._fields
    resolved = {}

    # Paso 1: por nombre técnico
    for key, candidates in PARTNER_FIELD_CANDIDATES.items():
        for c in candidates:
            if c in available:
                resolved[key] = c
                break

    # Paso 2: por etiqueta visible (cubre campos creados con Odoo Studio)
    for fname, fobj in available.items():
        label = (getattr(fobj, 'string', '') or '').lower()
        for key, patterns in LABEL_PATTERNS.items():
            if key not in resolved:
                if any(p in label for p in patterns):
                    # Verificar que sea un campo editable (no computed sin inverse)
                    if not getattr(fobj, 'compute', None) or getattr(fobj, 'inverse', None):
                        resolved[key] = fname
                        break

    return resolved


# ─────────────────────────────────────────────────────────────────
# LÍNEA DE MAPEO (TransientModel)
# ─────────────────────────────────────────────────────────────────
class LukerImportMappingLine(models.TransientModel):
    _name = 'luker.participant.import.mapping.line'
    _description = 'Línea de Mapeo de Columna'
    _order = 'sequence'

    wizard_id = fields.Many2one('luker.participant.import.wizard', required=True, ondelete='cascade')
    sequence = fields.Integer(default=10)

    column_name   = fields.Char(string='Columna del archivo', readonly=True)
    sample_value  = fields.Char(string='Dato ejemplo 1', readonly=True)
    sample_value2 = fields.Char(string='Dato ejemplo 2', readonly=True)
    sample_value3 = fields.Char(string='Dato ejemplo 3', readonly=True)
    auto_detected = fields.Boolean(readonly=True)

    destination_field = fields.Selection(
        selection=DESTINATION_FIELDS,
        string='Llevar a campo del sistema',
        required=True,
        default='_skip',
    )
    destination_label = fields.Char(compute='_compute_destination_label')
    is_skipped        = fields.Boolean(compute='_compute_is_skipped', store=True)
    status_icon       = fields.Char(compute='_compute_status_icon')

    @api.depends('destination_field')
    def _compute_destination_label(self):
        labels = dict(DESTINATION_FIELDS)
        for l in self:
            l.destination_label = labels.get(l.destination_field, '—')

    @api.depends('destination_field')
    def _compute_is_skipped(self):
        for l in self:
            l.is_skipped = (l.destination_field == '_skip')

    @api.depends('destination_field', 'auto_detected')
    def _compute_status_icon(self):
        for l in self:
            if l.destination_field == '_skip':
                l.status_icon = '⊘'
            elif l.auto_detected:
                l.status_icon = '✓'
            else:
                l.status_icon = '✎'


# ─────────────────────────────────────────────────────────────────
# WIZARD PRINCIPAL (TransientModel)
# ─────────────────────────────────────────────────────────────────
class LukerParticipantImportWizard(models.TransientModel):
    _name = 'luker.participant.import.wizard'
    _description = 'Carga Masiva de Participantes'

    step = fields.Selection([
        ('upload',   '1. Subir archivo'),
        ('mapping',  '2. Mapear columnas'),
        ('validate', '3. Validar'),
        ('done',     '4. Resultado'),
    ], default='upload', readonly=True)

    import_name = fields.Char(
        string='Nombre de la carga', required=True,
        default=lambda self: f'Carga SIMAT {datetime.now().strftime("%Y-%m-%d %H:%M")}',
    )
    source_format = fields.Selection([
        ('simat',  'SIMAT — Anexo 6A (formato oficial MEN)'),
        ('manual', 'Plantilla manual Luker'),
        ('otro',   'Otro formato Excel'),
        ('excel',  'Otro formato Excel (legacy)'),
    ], string='Formato del archivo', default='simat', required=True)

    file_data = fields.Binary(string='Archivo Excel (.xlsx)', required=True)
    file_name = fields.Char()

    reconcile_field = fields.Selection([
        ('doc_number',      'Número de documento'),
        ('doc_number_type', 'Documento + Tipo de documento'),
    ], string='Conciliar participantes por', default='doc_number', required=True)

    action_on_existing = fields.Selection([
        ('actualizar', 'Actualizar sus datos'),
        ('omitir',     'Omitir (no modificar)'),
        ('error',      'Marcar como error'),
    ], string='Si el participante ya existe', default='actualizar', required=True)

    participant_type_id = fields.Many2one(
        'luker.participant.type',
        string='Tipo de participante',
        required=True,
        default=lambda self: self.env['luker.participant.type'].search(
            [('cod_tipo_participante', '=', 'estudiante')], limit=1),
    )
    create_missing_contacts = fields.Boolean(string='Crear contacto si no existe', default=True)
    create_missing_orgs     = fields.Boolean(string='Crear organizaciones nuevas si no existen', default=True)

    mapping_line_ids = fields.One2many(
        'luker.participant.import.mapping.line', 'wizard_id', string='Mapeo de columnas')

    mapped_count  = fields.Integer(compute='_compute_mapping_stats')
    skipped_count = fields.Integer(compute='_compute_mapping_stats')
    auto_count    = fields.Integer(compute='_compute_mapping_stats')
    total_columns = fields.Integer(compute='_compute_mapping_stats')

    @api.depends('mapping_line_ids', 'mapping_line_ids.destination_field', 'mapping_line_ids.auto_detected')
    def _compute_mapping_stats(self):
        for wiz in self:
            lines = wiz.mapping_line_ids
            wiz.total_columns = len(lines)
            wiz.mapped_count  = len(lines.filtered(lambda l: l.destination_field != '_skip'))
            wiz.skipped_count = len(lines.filtered(lambda l: l.destination_field == '_skip'))
            wiz.auto_count    = len(lines.filtered(lambda l: l.auto_detected and l.destination_field != '_skip'))

    preview_html            = fields.Html(readonly=True)
    validation_warnings     = fields.Html(readonly=True)
    total_rows_preview      = fields.Integer(readonly=True)
    empty_alerts_count      = fields.Integer(readonly=True)
    duplicate_preview_count = fields.Integer(readonly=True)
    log_id         = fields.Many2one('luker.participant.import.log', readonly=True)
    result_summary = fields.Html(readonly=True)

    @api.onchange('source_format')
    def _onchange_source_format(self):
        if self.source_format == 'simat':
            tipo = self.env['luker.participant.type'].search([('cod_tipo_participante', '=', 'estudiante')], limit=1)
            if tipo:
                self.participant_type_id = tipo

    # ─────────────────────────────────────────────────────────
    # PASO 1 → 2
    # ─────────────────────────────────────────────────────────
    def action_analyze_file(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_('Por favor sube un archivo Excel antes de continuar.'))
        try:
            import openpyxl
        except ImportError:
            raise UserError(_('Se requiere openpyxl: pip install openpyxl'))

        raw = base64.b64decode(self.file_data)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        all_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            all_rows.append(row)
            if i >= 4: break
        wb.close()

        if not all_rows:
            raise UserError(_('El archivo está vacío.'))
        headers = all_rows[0]
        sample_rows = all_rows[1:4]
        if not any(headers):
            raise UserError(_('No se encontraron encabezados en la primera fila.'))

        self.mapping_line_ids.unlink()
        lines = []
        for seq, col in enumerate(headers):
            if col is None:
                continue
            col_str = str(col).strip()
            dest = SIMAT_AUTO_MAP.get(col_str, '_skip')
            auto = dest != '_skip'
            samples = []
            for sr in sample_rows:
                if seq < len(sr) and sr[seq] is not None:
                    v = str(sr[seq]).strip()
                    if v and v not in ('None', 'NO APLICA', '0', '99', '9'):
                        samples.append(v[:50])
            lines.append({
                'wizard_id': self.id, 'sequence': seq + 1,
                'column_name': col_str,
                'sample_value':  samples[0] if len(samples) > 0 else '—',
                'sample_value2': samples[1] if len(samples) > 1 else '',
                'sample_value3': samples[2] if len(samples) > 2 else '',
                'destination_field': dest,
                'auto_detected': auto,
            })
        self.env['luker.participant.import.mapping.line'].create(lines)
        self.step = 'mapping'
        return self._reopen()

    # ─────────────────────────────────────────────────────────
    # PASO 2 → 3
    # ─────────────────────────────────────────────────────────
    def action_validate(self):
        self.ensure_one()
        mapping = {l.column_name: l.destination_field for l in self.mapping_line_ids}
        dest_to_cols = {}
        for col, dest in mapping.items():
            if dest != '_skip':
                dest_to_cols.setdefault(dest, []).append(col)

        warnings_html = ''
        if 'doc_number' not in dest_to_cols:
            warnings_html += (
                '<div class="alert alert-danger"><b>Crítico:</b> '
                'Ninguna columna mapeada a "Número de documento". '
                'Sin este campo no se puede importar.</div>')

        raw = base64.b64decode(self.file_data)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        detected = [l.column_name for l in self.mapping_line_ids]
        rows_data = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue
            rows_data.append(dict(zip(detected, row)))
        wb.close()

        total = len(rows_data)
        empty_by_field = {}
        all_docs = {}
        field_labels = {'doc_number': 'Número de documento', 'first_name': 'Primer nombre',
                        'first_surname': 'Primer apellido', 'grade': 'Grado',
                        'institution_name': 'Nombre institución'}

        for row in rows_data:
            for imp in IMPORTANT_COLUMNS:
                cols = dest_to_cols.get(imp, [])
                if not any(row.get(c) for c in cols):
                    empty_by_field[imp] = empty_by_field.get(imp, 0) + 1
            doc_cols = dest_to_cols.get('doc_number', [])
            doc_val  = next((str(row.get(c,'')).strip() for c in doc_cols if row.get(c)), None)
            if doc_val:
                all_docs[doc_val] = all_docs.get(doc_val, 0) + 1

        dups_in_file  = sum(1 for v in all_docs.values() if v > 1)
        existing_in_db = self.env['luker.participant.identity'].search_count(
            [('num_identidad', 'in', list(all_docs.keys()))])

        alerts = []
        for field, count in empty_by_field.items():
            if count > 0:
                pct  = round(count * 100 / total, 1) if total else 0
                icon = '🔴' if field == 'doc_number' else '🟡'
                alerts.append(f'<li>{icon} <b>{field_labels.get(field,field)}</b>: {count:,} vacías ({pct}%)</li>')
        if dups_in_file:
            alerts.append(f'<li>🟠 <b>Documentos duplicados en el archivo</b>: {dups_in_file:,}</li>')
        if existing_in_db:
            al = dict(self._fields['action_on_existing'].selection).get(self.action_on_existing,'')
            alerts.append(f'<li>🔵 <b>Ya existen en BD</b>: {existing_in_db:,} — acción: <b>{al}</b></li>')

        if alerts:
            warnings_html += ('<div class="alert alert-warning"><h6>⚠ Alertas</h6>'
                              '<ul style="margin-bottom:0;">' + ''.join(alerts) + '</ul></div>')
        else:
            warnings_html = '<div class="alert alert-success">✓ Sin alertas críticas.</div>'

        mapped_lines = [(l.column_name, l.destination_field, l.destination_label)
                        for l in self.mapping_line_ids if l.destination_field != '_skip'][:8]
        header = ''.join(
            f'<th style="background:#f7f0f5;color:#4e3348;padding:7px 10px;border:1px solid #ddd;'
            f'font-size:11px;white-space:nowrap;">{dl.split("→")[0].strip()}'
            f'<br><small style="color:#999;font-style:italic;">{col}</small></th>'
            for col, _, dl in mapped_lines)
        body = ''
        for row in rows_data[:8]:
            cells = ''
            for col, _, _ in mapped_lines:
                val = row.get(col,'')
                if val is None or str(val).strip() in ('','None','NO APLICA'):
                    cells += '<td style="padding:5px 10px;border:1px solid #eee;font-size:11px;color:#dc3545;">—</td>'
                else:
                    cells += f'<td style="padding:5px 10px;border:1px solid #eee;font-size:11px;">{str(val)[:35]}</td>'
            body += f'<tr>{cells}</tr>'

        self.validation_warnings     = warnings_html
        self.preview_html            = (
            f'<p style="font-size:12px;color:#714B67;font-weight:600;">'
            f'Vista previa — 8 primeras filas · {len(mapped_lines)} columnas mapeadas</p>'
            f'<div style="overflow-x:auto;"><table style="border-collapse:collapse;width:100%;">'
            f'<thead><tr>{header}</tr></thead><tbody>{body}</tbody></table></div>')
        self.total_rows_preview      = total
        self.empty_alerts_count      = sum(1 for v in empty_by_field.values() if v > 0)
        self.duplicate_preview_count = dups_in_file + existing_in_db
        self.step = 'validate'
        return self._reopen()

    # ─────────────────────────────────────────────────────────
    # PASO 3 → 4 — IMPORTAR
    # ─────────────────────────────────────────────────────────
    def action_import(self):
        self.ensure_one()
        mapping  = {l.column_name: l.destination_field for l in self.mapping_line_ids}
        detected = [l.column_name for l in self.mapping_line_ids]
        dest_to_cols = {}
        for col, dest in mapping.items():
            if dest != '_skip':
                dest_to_cols.setdefault(dest, []).append(col)

        raw = base64.b64decode(self.file_data)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        rows_data = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0: continue
            rows_data.append(dict(zip(detected, row)))
        wb.close()

        # Precargar tipos de identificación
        all_id_types   = self.env['l10n_latam.identification.type'].search([])
        id_type_by_name = {t.name.lower(): t for t in all_id_types}

        # Resolver campos disponibles en res.partner
        pf = _resolve_partner_fields(self.env['res.partner'])
        _logger.info(
            'Luker Import — campos detectados en res.partner: %s',
            {k: v for k, v in pf.items()}
        )
        _logger.info(
            'Luker Import — campos NO encontrados: %s',
            [k for k in ['first_name','second_name','first_surname','second_surname',
                          'estrato','sisben_bool','sisben_class','doc_number']
             if k not in pf]
        )

        skipped_cols = [l.column_name for l in self.mapping_line_ids if l.destination_field == '_skip']
        mapping_summary = {l.column_name: l.destination_label
                           for l in self.mapping_line_ids if l.destination_field != '_skip'}

        log = self.env['luker.participant.import.log'].create({
            'nom_carga':              self.import_name,
            'nom_archivo_origen':     self.file_name,
            'archivo_data':           self.file_data,
            'fuente_formato':         self.source_format or 'otro',
            'total_filas':            len(rows_data),
            'campo_conciliacion':     self.reconcile_field,
            'accion_existente':       self.action_on_existing,
            'mapeo_columnas_json':    json.dumps(mapping_summary, ensure_ascii=False, indent=2),
            'columnas_omitidas_json': json.dumps(skipped_cols, ensure_ascii=False),
            'estado_carga':           'borrador',
        })

        counters  = {'creado': 0, 'actualizado': 0, 'omitido': 0, 'error': 0, 'advertencia': 0}
        log_lines = []

        def _get(row, field):
            for c in dest_to_cols.get(field, []):
                v = row.get(c)
                if v is not None and str(v).strip() not in ('', 'None', 'NO APLICA', '99', '0'):
                    return v
            return None

        def _find_id_type(code, name_raw):
            search = SIMAT_DOC_TYPE_NAMES.get(code)
            if not search and name_raw:
                search = str(name_raw).strip()
            if search:
                t = id_type_by_name.get(search.lower())
                if t: return t
                for k, t in id_type_by_name.items():
                    if search.lower() in k or k in search.lower():
                        return t
            return all_id_types[:1] and all_id_types[0] or False

        def _build_partner_vals(row, full_name, first_name, second_name,
                                first_surname, second_surname,
                                phone_raw, birthdate, gender_raw, pf):
            """Construye los valores del contacto mapeando todos los campos disponibles."""
            vals = {
                'name': full_name,
                'is_company': False,
            }

            # ── Nombre separado ──────────────────────────────
            # También aseguramos que el campo 'name' tenga el nombre completo
            for key, val in [('first_name', first_name), ('second_name', second_name),
                              ('first_surname', first_surname), ('second_surname', second_surname)]:
                f = pf.get(key)
                if f and val:
                    vals[f] = val

            # ── Teléfono ─────────────────────────────────────
            if phone_raw:
                phone_str = str(phone_raw).strip()
                if pf.get('phone') and phone_str:
                    vals[pf['phone']] = phone_str

            # ── Fecha de nacimiento ──────────────────────────
            if birthdate and pf.get('birthdate'):
                vals[pf['birthdate']] = birthdate

            # ── Género ───────────────────────────────────────
            gender_raw_up = str(gender_raw or '').strip().upper()
            if gender_raw_up and pf.get('gender_bio'):
                vals[pf['gender_bio']] = GENDER_MAP.get(gender_raw_up, gender_raw_up.lower())
            if gender_raw_up and pf.get('gender_id'):
                vals[pf['gender_id']] = GENDER_MAP.get(gender_raw_up, gender_raw_up.lower())

            # ── Dirección ─────────────────────────────────────
            addr = _get(row, 'address')
            if addr and pf.get('street'):
                vals[pf['street']] = str(addr)

            # ── Zona ─────────────────────────────────────────
            zone_raw = _get(row, 'zone')
            if zone_raw is not None:
                zone_val = ZONE_MAP.get(zone_raw, ZONE_MAP.get(str(zone_raw).upper(), str(zone_raw)))
                if pf.get('zone'):
                    vals[pf['zone']] = zone_val

            # ── Estrato ──────────────────────────────────────
            estrato = _get(row, 'estrato')
            if estrato is not None and pf.get('estrato'):
                try:
                    vals[pf['estrato']] = int(str(estrato).strip())
                except (ValueError, TypeError):
                    pass

            # ── SISBEN ───────────────────────────────────────
            sisben = _get(row, 'sisben')
            if sisben:
                sisben_str = str(sisben).strip()
                if pf.get('sisben_bool'):
                    vals[pf['sisben_bool']] = True
                if pf.get('sisben_class'):
                    vals[pf['sisben_class']] = sisben_str

            # ── Discapacidad ─────────────────────────────────
            disc_code = _get(row, 'disability_code')
            disc_name = _get(row, 'disability_name')
            has_disability = disc_code is not None and str(disc_code).strip() not in ('99', '0', '9', '')
            if has_disability:
                if pf.get('disability'):
                    vals[pf['disability']] = True
                if disc_name and pf.get('disability_type'):
                    vals[pf['disability_type']] = str(disc_name).strip()

            # ── Víctima de conflicto ──────────────────────────
            victim = str(_get(row, 'victim_conflict') or '').strip().upper()
            is_victim = victim in ('SI', 'S', '1', 'YES', 'Y')
            if is_victim and pf.get('victim'):
                vals[pf['victim']] = True

            # ── Lugar de expedición ───────────────────────────
            exp_depto = _get(row, 'exp_depto')
            if exp_depto and pf.get('exp_place'):
                vals[pf['exp_place']] = str(exp_depto)

            return vals

        for row_num, row in enumerate(rows_data, start=2):
            result = 'creado'
            empty_fields = []

            try:
                sp = self.env.cr.savepoint()
                sp.__enter__()
            except Exception:
                sp = None

            try:
                doc_number      = _get(row, 'doc_number')
                doc_type_code   = _get(row, 'doc_type_code')
                doc_type_name_r = _get(row, 'doc_type_name')
                first_name      = str(_get(row, 'first_name')    or '').strip().title()
                second_name     = str(_get(row, 'second_name')   or '').strip().title()
                first_surname   = str(_get(row, 'first_surname') or '').strip().title()
                second_surname  = str(_get(row, 'second_surname')or '').strip().title()
                birthdate_raw   = _get(row, 'birthdate')
                gender_raw      = str(_get(row, 'gender') or '').strip().upper()
                phone_raw       = _get(row, 'phone')
                grade_raw       = _get(row, 'grade')
                group_raw       = _get(row, 'group_name')
                jornada_code    = _get(row, 'jornada_code')
                jornada_name    = str(_get(row, 'jornada_name') or '').strip().upper()
                inst_name       = str(_get(row, 'institution_name') or '').strip()
                sede_name       = str(_get(row, 'sede_name') or inst_name).strip()
                estrato         = _get(row, 'estrato')
                sisben          = _get(row, 'sisben')
                victim          = str(_get(row, 'victim_conflict') or '').strip().upper()
                repeating       = str(_get(row, 'repeating') or '').strip().upper()

                if not first_name:    empty_fields.append('Primer nombre')
                if not first_surname: empty_fields.append('Primer apellido')
                if not grade_raw:     empty_fields.append('Grado')
                if not inst_name:     empty_fields.append('Institución')

                if not doc_number:
                    counters['advertencia'] += 1
                    log_lines.append({'carga_id': log.id, 'num_fila': row_num,
                                      'estado_registro': 'advertencia',
                                      'mensaje_validacion': 'Sin número de documento — omitida',
                                      'campos_vacios_json': ', '.join(['Nº documento'] + empty_fields)})
                    continue

                full_name = ' '.join(filter(None, [first_name, second_name, first_surname, second_surname]))
                if not full_name.strip():
                    full_name = f'Participante {doc_number}'

                # Fecha de nacimiento
                birthdate = None
                if birthdate_raw:
                    if isinstance(birthdate_raw, datetime):
                        birthdate = birthdate_raw.date()
                    elif hasattr(birthdate_raw, 'date'):
                        birthdate = birthdate_raw.date()
                    elif isinstance(birthdate_raw, str):
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                            try:
                                birthdate = datetime.strptime(birthdate_raw.strip(), fmt).date()
                                break
                            except ValueError:
                                pass

                gender_mapped = GENDER_MAP.get(gender_raw, False)

                # Conciliar
                existing_identity = self.env['luker.participant.identity'].search([
                    ('num_identidad', '=', str(doc_number).strip()), ('es_principal', '=', True)], limit=1)

                if existing_identity and self.action_on_existing in ('omitir', 'skip'):
                    counters['omitido'] += 1
                    log_lines.append({'carga_id': log.id, 'num_fila': row_num,
                                      'estado_registro': 'omitido',
                                      'participante_id': existing_identity.participant_id.id,
                                      'mensaje_validacion': f'Ya existe — omitido (doc: {doc_number})',
                                      'campos_vacios_json': ', '.join(empty_fields)})
                    continue

                if existing_identity and self.action_on_existing == 'error':
                    counters['error'] += 1
                    log_lines.append({'carga_id': log.id, 'num_fila': row_num,
                                      'estado_registro': 'error',
                                      'mensaje_validacion': f'Ya existe y acción es "error" (doc: {doc_number})'})
                    continue

                # Construir vals del partner
                partner_vals = _build_partner_vals(
                    row, full_name, first_name, second_name,
                    first_surname, second_surname, phone_raw, birthdate, gender_raw, pf)

                if existing_identity:
                    # Actualizar partner existente
                    partner = existing_identity.participant_id.partner_id
                    if partner:
                        partner.write({k: v for k, v in partner_vals.items() if k != 'name' and k != 'is_company'})
                    # Actualizar participante
                    p_write = {}
                    if birthdate:      p_write['fecha_nacimiento'] = birthdate
                    if gender_mapped:  p_write['sexo']          = gender_mapped
                    if p_write:
                        existing_identity.participant_id.write(p_write)
                    participant = existing_identity.participant_id
                    result = 'actualizado'
                    counters['actualizado'] += 1
                else:
                    # Crear partner
                    if self.create_missing_contacts:
                        partner = self.env['res.partner'].search([
                            ('name', '=ilike', full_name), ('is_company', '=', False)], limit=1)
                        if not partner:
                            # Crear con campos básicos seguros primero
                            base_vals = {'name': partner_vals['name'], 'is_company': False}
                            if partner_vals.get('phone'):
                                base_vals['phone'] = partner_vals['phone']
                            partner = self.env['res.partner'].create(base_vals)
                            # Luego escribir campos adicionales uno por uno
                            for k, v in partner_vals.items():
                                if k in ('name', 'is_company', 'phone'):
                                    continue
                                try:
                                    with self.env.cr.savepoint():
                                        partner.write({k: v})
                                except Exception as ef:
                                    _logger.warning('Campo partner.%s no pudo escribirse: %s', k, ef)
                        else:
                            safe_update = {k: v for k, v in partner_vals.items()
                                           if k not in ('name', 'is_company')}
                            for k, v in safe_update.items():
                                try:
                                    with self.env.cr.savepoint():
                                        partner.write({k: v})
                                except Exception as ew:
                                    _logger.warning('No se pudo actualizar partner.%s: %s', k, ew)
                    else:
                        partner = self.env['res.partner'].search([
                            ('name', '=ilike', full_name), ('is_company', '=', False)], limit=1)
                        if not partner:
                            counters['error'] += 1
                            log_lines.append({'carga_id': log.id, 'num_fila': row_num,
                                              'estado_registro': 'error',
                                              'mensaje_validacion': f'Contacto no encontrado: {full_name}'})
                            continue

                    participant = self.env['luker.participant'].create({
                        'partner_id':          partner.id,
                        'tipo_participante_id': self.participant_type_id.id,
                        'fecha_nacimiento':     birthdate,
                        'sexo':                 gender_mapped,
                        'estado':               'activo',
                        'carga_origen_id':       log.id,
                    })

                    # Identificación en luker
                    id_type = _find_id_type(doc_type_code, doc_type_name_r)
                    self.env['luker.participant.identity'].create({
                        'participante_id':        participant.id,
                        'tipo_identidad_id':  id_type.id if id_type else False,
                        'num_identidad':      str(doc_number).strip(),
                        'es_principal':       True,
                        'estado':             'activa',
                    })
                    # Tipo de identificación y número en el contacto
                    id_write_vals = {}
                    if id_type and pf.get('id_type'):
                        id_write_vals[pf['id_type']] = id_type.id
                    if doc_number:
                        doc_str = str(doc_number).strip()
                        # Intentar primero l10n_latam_document_number, luego vat
                        doc_field = pf.get('doc_number', 'vat')
                        id_write_vals[doc_field] = doc_str
                    if id_write_vals:
                        for k, v in id_write_vals.items():
                            try:
                                with self.env.cr.savepoint():
                                    partner.write({k: v})
                            except Exception as e_id:
                                _logger.warning('Campo id partner.%s: %s', k, e_id)

                    result = 'creado'
                    counters['creado'] += 1

                # Contexto organizacional
                if inst_name and grade_raw:
                    org = self.env['luker.organization'].search(
                        [('nom_unidad', '=ilike', inst_name)], limit=1)
                    if not org and self.create_missing_orgs:
                        partner_org = self.env['res.partner'].search([
                            ('name', '=ilike', inst_name), ('is_company', '=', True)], limit=1)
                        if not partner_org:
                            partner_org = self.env['res.partner'].create(
                                {'name': inst_name, 'is_company': True})
                        org = self.env['luker.organization'].create(
                            {'partner_id': partner_org.id, 'tipo_dominio': 'educacion_formal'})
                    if org:
                        branch = self.env['luker.organization.branch'].search([
                            ('institucion_id', '=', org.id),
                            ('nom_sede', '=ilike', sede_name or inst_name)], limit=1)
                        if not branch:
                            branch = self.env['luker.organization.branch'].create(
                                {'institucion_id': org.id, 'nom_sede': sede_name or 'Sede Principal'})
                        jornada = (JORNADA_MAP.get(jornada_code)
                                   or JORNADA_MAP.get(str(jornada_code), False)
                                   or JORNADA_MAP.get(jornada_name, False))
                        unit = self.env['luker.organization.unit'].search([
                            ('sede_id', '=', branch.id), ('nom_grado', '=', str(grade_raw)),
                            ('nom_grupo', '=', str(group_raw) if group_raw else False)], limit=1)
                        if not unit:
                            unit = self.env['luker.organization.unit'].create({
                                'sede_id': branch.id, 'nom_grado': str(grade_raw),
                                'nom_grupo': str(group_raw) if group_raw else False,
                                'jornada': jornada})
                        if not self.env['luker.participant.assignment'].search([
                                ('participante_id', '=', participant.id),
                                ('vigencia_hasta', '=', False)], limit=1):
                            self.env['luker.participant.assignment'].create({
                                'participante_id': participant.id, 'institucion_id': org.id,
                                'sede_id': branch.id, 'unidad_id': unit.id})

                # Atributos dinámicos
                attr_defs = {a.cod_atributo: a for a in self.env['luker.attribute.definition'].search([])}
                def _write_attr(code, val_field, raw_val):
                    if raw_val is None or (isinstance(raw_val, bool) and not raw_val):
                        return
                    attr_def = attr_defs.get(code)
                    if not attr_def:
                        return
                    if not self.env['luker.participant.attribute.value'].search([
                            ('participante_id', '=', participant.id),
                            ('definicion_id', '=', attr_def.id), ('vigencia_hasta', '=', False)], limit=1):
                        self.env['luker.participant.attribute.value'].create({
                            'participante_id': participant.id, 'definicion_id': attr_def.id,
                            val_field: raw_val})

                estrato_int = None
                if estrato is not None:
                    try: estrato_int = int(str(estrato).strip())
                    except: pass

                _write_attr('attr_estrato',        'valor_numero', estrato_int)
                _write_attr('attr_sisben',          'valor_texto',    str(sisben).strip() if sisben else None)
                _write_attr('attr_desplazamiento',  'valor_booleano', victim in ('SI','S','1','YES','Y') if victim else None)
                _write_attr('attr_repitente',       'valor_booleano', repeating in ('S','SI','1','Y','YES'))

                disc_code_raw = _get(row, 'disability_code')
                has_disc = disc_code_raw is not None and str(disc_code_raw).strip() not in ('99','0','9','')
                _write_attr('attr_discapacidad',    'valor_booleano', has_disc if has_disc else None)

                log_lines.append({'carga_id': log.id, 'num_fila': row_num,
                                   'estado_registro': result, 'participante_id': participant.id,
                                   'mensaje_validacion': f'{"Creado" if result == "creado" else "Actualizado"}: {full_name}',
                                   'campos_vacios_json': ', '.join(empty_fields) if empty_fields else ''})

            except Exception as e:
                counters['error'] += 1
                _logger.exception('Error fila %s: %s', row_num, e)
                # Revertir solo esta fila para no contaminar el resto de la transacción
                if sp is not None:
                    try:
                        sp.__exit__(type(e), e, e.__traceback__)
                    except Exception:
                        pass
                    sp = None
                log_lines.append({'carga_id': log.id, 'num_fila': row_num,
                                   'estado_registro': 'error', 'mensaje_validacion': str(e)[:200]})
                continue
            else:
                if sp is not None:
                    try:
                        sp.__exit__(None, None, None)
                    except Exception:
                        pass

        if log_lines:
            self.env['luker.participant.import.log.line'].create(log_lines)

        if not counters['error']:
            estado_carga_final = 'completo'
        elif counters['creado'] + counters['actualizado'] > 0:
            estado_carga_final = 'parcial'
        else:
            estado_carga_final = 'fallido'

        # Actualizar OPE_Carga_Poblacion usando SQL directo para evitar problemas de caché
        self.env.cr.execute(
            """UPDATE luker_participant_import_log
               SET estado_carga    = %s,
                   filas_validas   = %s,
                   filas_omitidas  = %s,
                   filas_invalidas = %s,
                   alertas_vacios  = %s
               WHERE id = %s""",
            (
                estado_carga_final,
                counters['creado'] + counters['actualizado'],
                counters['omitido'],
                counters['error'],
                counters['advertencia'],
                log.id,
            )
        )
        log.invalidate_recordset()

        self.result_summary = (
            f'<div style="font-family:sans-serif;"><h4 style="color:#714B67;">Importación completada</h4>'
            f'<table style="border-collapse:collapse;width:100%;max-width:400px;">'
            f'<tr style="background:#d4edda;"><td style="padding:6px 14px;border:1px solid #c3e6cb;">✓ Creados</td>'
            f'<td style="padding:6px 14px;border:1px solid #c3e6cb;font-weight:bold;text-align:right;">{counters["creado"]:,}</td></tr>'
            f'<tr style="background:#d1ecf1;"><td style="padding:6px 14px;border:1px solid #bee5eb;">↻ Actualizados</td>'
            f'<td style="padding:6px 14px;border:1px solid #bee5eb;font-weight:bold;text-align:right;">{counters["actualizado"]:,}</td></tr>'
            f'<tr style="background:#fff3cd;"><td style="padding:6px 14px;border:1px solid #ffeeba;">○ Omitidos</td>'
            f'<td style="padding:6px 14px;border:1px solid #ffeeba;text-align:right;">{counters["omitido"]:,}</td></tr>'
            f'<tr style="background:#f8d7da;"><td style="padding:6px 14px;border:1px solid #f5c6cb;">✗ Errores</td>'
            f'<td style="padding:6px 14px;border:1px solid #f5c6cb;text-align:right;">{counters["error"]:,}</td></tr>'
            f'</table><p style="font-size:12px;color:#666;margin-top:10px;">'
            f'Log: <b>{log.nom_carga}</b> · Por: <b>{self.env.user.name}</b></p></div>')
        self.log_id = log.id
        self.step   = 'done'
        return self._reopen()

    def action_back_to_upload(self):  self.step = 'upload';   return self._reopen()
    def action_back_to_mapping(self): self.step = 'mapping';  return self._reopen()
    def action_back_to_validate(self):self.step = 'validate'; return self._reopen()

    def action_view_log(self):
        return {'type': 'ir.actions.act_window', 'name': 'Log de importación',
                'res_model': 'luker.participant.import.log',
                'res_id': self.log_id.id, 'view_mode': 'form'}

    def action_open_import_wizard(self):
        return {'type': 'ir.actions.act_window', 'name': 'Carga Masiva de Participantes',
                'res_model': 'luker.participant.import.wizard',
                'view_mode': 'form', 'target': 'new'}

    def _reopen(self):
        return {'type': 'ir.actions.act_window', 'res_model': self._name,
                'res_id': self.id, 'view_mode': 'form', 'target': 'new'}
