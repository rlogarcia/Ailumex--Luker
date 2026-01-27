# Modelo para importar leads y contactos desde archivos CSV o Excel
# Comentarios y docstring en español para facilitar el mantenimiento.

import base64
import csv
import io
import logging
import os
from odoo import _, fields, models, api
from odoo.exceptions import UserError

# Librerías para leer archivos Excel
try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

_logger = logging.getLogger(__name__)


class ImportLeadsWizard(models.Model):
    """
    Modelo para importar leads y contactos desde archivos CSV o Excel.
    Permite descargar la plantilla, subir el archivo y mostrar el resultado.
    """

    _name = "import.leads.wizard"
    _description = "Importar leads y contactos desde archivos CSV o Excel"
    _order = "create_date desc"
    _inherit = ["mail.thread", "mail.activity.mixin"]

    name = fields.Char(string="Referencia", required=True, copy=False, tracking=True)
    file_data = fields.Binary(string="Archivo", required=False, attachment=False)
    file_name = fields.Char(string="Nombre del archivo", tracking=True)
    result_message = fields.Text(string="Resultado", readonly=True)
    scoring_details = fields.Html(
        string="Detalles de Scoring",
        readonly=True,
        sanitize=False,
        help="Tabla HTML con detalles del puntaje calculado para cada lead",
    )
    enrichment_details = fields.Html(
        string="Detalles de Enriquecimiento",
        readonly=True,
        sanitize=False,
        help="Tabla HTML con información enriquecida de empresas",
    )
    state = fields.Selection(
        [
            ("draft", "Borrador"),
            ("done", "Completado"),
        ],
        string="Estado",
        default="draft",
        readonly=True,
        copy=False,
        tracking=True,
    )
    leads_count = fields.Integer(
        string="Leads Nuevos", readonly=True, default=0, copy=False
    )
    leads_updated = fields.Integer(
        string="Leads Actualizados", readonly=True, default=0, copy=False
    )
    contacts_created = fields.Integer(
        string="Contactos Creados", readonly=True, default=0, copy=False
    )
    contacts_updated = fields.Integer(
        string="Contactos Actualizados", readonly=True, default=0, copy=False
    )
    lead_ids = fields.Many2many(
        "crm.lead", string="Leads Creados", readonly=True, copy=False
    )

    @api.model
    def default_get(self, fields_list):
        """Genera un nombre automático para nuevas importaciones"""
        res = super(ImportLeadsWizard, self).default_get(fields_list)
        if "name" in fields_list and not res.get("name"):
            # Contar cuántas importaciones hay
            count = self.search_count([]) + 1
            res["name"] = f"Importación #{count:04d}"
        return res

    def action_view_leads(self):
        """Acción para ver los leads creados en esta importación"""
        self.ensure_one()
        return {
            "name": _("Oportunidades Importadas"),
            "type": "ir.actions.act_window",
            "res_model": "crm.lead",
            "view_mode": "kanban,list,form",
            "domain": [("id", "in", self.lead_ids.ids)],
            "context": {"default_type": "opportunity"},
        }

    # Encabezados obligatorios para la importación
    _required_headers = [
        "Nombre",
        "Correo",
        "Telefono",
        "Teléfono 2",
        "Empresa",
        "Pais",
        "Ciudad",
        "Fuente",
        "Marca campaña",
        "Nombre campaña",
        "Curso / Programa interés",
        "Perfil",
        "Observaciones",
        "Monto",
    ]

    def action_download_template(self):
        # Acción para descargar la plantilla de importación
        self.ensure_one()
        return {
            "type": "ir.actions.act_url",
            "url": "/crm/import_leads/template",
            "target": "self",
        }

    def _validate_row(self, row, line_number):
        """
        Valida que una fila tenga los datos mínimos requeridos.
        Retorna una lista de errores encontrados.
        """
        errors = []

        # Validar nombre (obligatorio)
        if not row.get("Nombre", "").strip():
            errors.append(f"Línea {line_number}: ⚠️ Falta el nombre del contacto")

        # Validar que tenga al menos un método de contacto (email o teléfono)
        email = row.get("Correo", "").strip()
        phone = row.get("Telefono", "").strip()
        phone_2 = row.get("Teléfono 2", "").strip()

        if not email and not phone and not phone_2:
            errors.append(
                f"Línea {line_number}: ⚠️ Debe tener al menos correo o teléfono"
            )

        # Validar formato de email si existe
        if email and "@" not in email:
            errors.append(
                f"Línea {line_number}: ⚠️ El correo '{email}' no tiene formato válido"
            )

        # Validar monto (debe ser numérico si existe)
        monto = row.get("Monto", "").strip()
        if monto:
            try:
                float(monto)
            except ValueError:
                errors.append(
                    f"Línea {line_number}: ⚠️ El monto '{monto}' no es un número válido"
                )

        return errors

    def _find_existing_lead(self, row):
        """
        Busca si ya existe un lead con los mismos datos.
        Retorna el ID del lead existente si lo encuentra, None si no.
        """
        email = row.get("Correo", "").strip()
        phone = row.get("Telefono", "").strip()
        phone_2 = row.get("Teléfono 2", "").strip()
        name = row.get("Nombre", "").strip()

        existing_lead_id = None

        # 1. Buscar por email (más confiable)
        if email:
            self.env.cr.execute(
                "SELECT id FROM crm_lead WHERE LOWER(email_from) = LOWER(%s) AND type = 'opportunity' ORDER BY write_date DESC LIMIT 1",
                (email,),
            )
            result = self.env.cr.fetchone()
            if result:
                existing_lead_id = result[0]
                _logger.info(
                    f"Lead existente encontrado por email {email}: ID {existing_lead_id}"
                )
                return existing_lead_id

        # 2. Buscar por teléfono (si no hay email o no se encontró por email)
        if phone and not existing_lead_id:
            self.env.cr.execute(
                "SELECT id FROM crm_lead WHERE phone = %s AND type = 'opportunity' ORDER BY write_date DESC LIMIT 1",
                (phone,),
            )
            result = self.env.cr.fetchone()
            if result:
                existing_lead_id = result[0]
                _logger.info(
                    f"Lead existente encontrado por teléfono {phone}: ID {existing_lead_id}"
                )
                return existing_lead_id

        # 2.1 Buscar por teléfono alterno si no se encontró por el principal
        if phone_2 and not existing_lead_id:
            self.env.cr.execute(
                "SELECT id FROM crm_lead WHERE (phone = %s OR mobile = %s) AND type = 'opportunity' ORDER BY write_date DESC LIMIT 1",
                (phone_2, phone_2),
            )
            result = self.env.cr.fetchone()
            if result:
                existing_lead_id = result[0]
                _logger.info(
                    f"Lead existente encontrado por teléfono 2 {phone_2}: ID {existing_lead_id}"
                )
                return existing_lead_id

        # 3. Buscar por nombre del contacto (menos confiable, solo si no hay email ni teléfono)
        if name and not existing_lead_id and not email and not phone:
            self.env.cr.execute(
                "SELECT id FROM crm_lead WHERE LOWER(contact_name) = LOWER(%s) AND type = 'opportunity' ORDER BY write_date DESC LIMIT 1",
                (name,),
            )
            result = self.env.cr.fetchone()
            if result:
                existing_lead_id = result[0]
                _logger.info(
                    f"Lead existente encontrado por nombre {name}: ID {existing_lead_id}"
                )
                return existing_lead_id

        return None

    def _update_existing_lead(self, lead_id, row, partner_id):
        """
        Actualiza un lead existente con nuevos datos del CSV.
        Solo actualiza campos que tienen valor en el CSV.
        """
        lead = self.env["crm.lead"].browse(lead_id)

        name = row.get("Nombre", "").strip()
        email = row.get("Correo", "").strip()
        phone = row.get("Telefono", "").strip()
        phone_2 = row.get("Teléfono 2", "").strip()
        company = row.get("Empresa", "").strip()
        country_name = row.get("Pais", "").strip()
        city = row.get("Ciudad", "").strip()
        source_name = row.get("Fuente", "").strip()
        medium_name = row.get("Marca campaña", "").strip()
        campaign_name = row.get("Nombre campaña", "").strip()
        program_interest = row.get("Curso / Programa interés", "").strip()
        profile = row.get("Perfil", "").strip()
        observations = row.get("Observaciones", "").strip()
        monto = row.get("Monto", "").strip()

        vals = {}

        if monto:
            try:
                vals["expected_revenue"] = float(monto)
            except ValueError:
                pass

        if country_name:
            country_id = self._find_country(country_name)
            if country_id:
                vals["country_id"] = country_id

        if city:
            vals["city"] = city

        if source_name:
            vals["source_id"] = self._get_or_create_utm_source(source_name)
        if medium_name:
            vals["medium_id"] = self._get_or_create_utm_medium(medium_name)
        if campaign_name:
            vals["campaign_id"] = self._get_or_create_utm_campaign(campaign_name)

        if phone:
            vals["phone"] = phone
        if phone_2:
            vals["mobile"] = phone_2
        if email:
            vals["email_from"] = email
        if partner_id:
            vals["partner_id"] = partner_id

        lead_name = company or name
        if lead_name:
            vals["name"] = lead_name
        if name:
            vals["contact_name"] = name

        if program_interest:
            vals["program_interest"] = program_interest
        if profile:
            mapped_profile = self._map_profile(profile)
            if mapped_profile:
                vals["profile"] = mapped_profile
        if observations:
            vals["description"] = observations

        if vals:
            lead.write(vals)
            _logger.info("Lead %s actualizado con nuevos datos del CSV", lead_id)
            return True

        return False

    def action_import(self):
        # Acción principal para importar los leads desde el archivo subido
        self.ensure_one()
        if not self.file_data:
            raise UserError(
                _("Por favor, selecciona un archivo CSV o Excel para importar.")
            )

        file_content = base64.b64decode(self.file_data)
        filename = self.file_name or ""
        extension = os.path.splitext(filename)[1].lower() if filename else ""
        if not extension:
            extension = self._detect_extension(file_content)

        if extension not in (".csv", ".xls", ".xlsx"):
            raise UserError(_("Formato no soportado. Usa archivos .csv, .xls o .xlsx."))

        # Leer las filas del archivo según el tipo
        rows = self._read_rows(file_content, extension)
        if not rows:
            raise UserError(_("El archivo está vacío o no contiene datos válidos."))

        created = 0
        created_contacts = 0
        updated_contacts = 0
        updated_leads = 0  # Nueva variable para contar leads actualizados
        skipped = []
        enriched_count = 0
        enrichment_log = []  # Lista para guardar datos enriquecidos
        lead_ids_created = []

        # Obtener servicio de enriquecimiento
        enrichment_service = self.env[
            "company.enrichment.service"
        ]._get_active_service()

        # Procesar cada fila del archivo
        for line_number, row in enumerate(rows, start=2):
            try:
                # 1️⃣ VALIDAR DATOS DE LA FILA
                validation_errors = self._validate_row(row, line_number)
                if validation_errors:
                    for error in validation_errors:
                        skipped.append(error)
                    continue  # Saltar esta fila si tiene errores de validación

                # 2️⃣ ENRIQUECER DATOS (si hay correo empresarial)
                email = row.get("Correo", "").strip()
                enriched_data = {}
                if email and enrichment_service:
                    enriched_data = enrichment_service.enrich_from_email(email)

                    if enriched_data:
                        # Guardar log de enriquecimiento con información completa del lead
                        enrichment_log.append(
                            {
                                "nombre": row.get("Nombre", ""),
                                "email": email,
                                "telefono": row.get("Telefono")
                                or row.get("Teléfono 2")
                                or "No proporcionado",
                                "enriched_data": enriched_data,
                                "line_number": line_number,
                            }
                        )

                        # Aplicar datos enriquecidos si no existen en el CSV
                        if "country" in enriched_data and not row.get("Pais"):
                            row["Pais"] = enriched_data["country"]
                            enriched_count += 1
                            _logger.info(
                                f'Línea {line_number}: País enriquecido → {enriched_data["country"]}'
                            )

                        if "company_name" in enriched_data and not row.get("Empresa"):
                            row["Empresa"] = enriched_data["company_name"]
                            _logger.info(
                                f'Línea {line_number}: Empresa enriquecida → {enriched_data["company_name"]}'
                            )

                # 3️⃣ VERIFICAR SI YA EXISTE EL LEAD
                existing_lead_id = self._find_existing_lead(row)

                with self.env.cr.savepoint():
                    partner_id, partner_stats = self._get_or_create_partner_sql(row)
                    updated_contacts += partner_stats.get("updated", 0)
                    created_contacts += partner_stats.get("created", 0)

                    if existing_lead_id:
                        # 4️⃣ ACTUALIZAR LEAD EXISTENTE
                        updated = self._update_existing_lead(
                            existing_lead_id, row, partner_id
                        )
                        if updated:
                            updated_leads += 1
                            lead_ids_created.append(
                                existing_lead_id
                            )  # Agregar para recalcular score
                            _logger.info(
                                f"Línea {line_number}: Lead existente actualizado (ID: {existing_lead_id})"
                            )
                    else:
                        # 5️⃣ CREAR NUEVO LEAD
                        lead_id = self._create_lead_sql(row, partner_id)
                        if lead_id:
                            created += 1
                            lead_ids_created.append(lead_id)
                            _logger.info(
                                f"Línea {line_number}: Nuevo lead creado (ID: {lead_id})"
                            )

            except UserError as error:
                skipped.append(
                    _(
                        "Línea %(line)s: ❌ %(error)s",
                        line=line_number,
                        error=str(error),
                    )
                )
                continue
            except Exception as error:
                _logger.exception("Error importing lead on line %s", line_number)
                skipped.append(
                    _(
                        "Línea %(line)s: ❌ Error inesperado: %(error)s",
                        line=line_number,
                        error=str(error),
                    )
                )
                continue

        # Construir el resumen del resultado
        summary = _(
            "✅ %(created)s lead(s) nuevos creados.\n"
            "🔄 %(updated_leads)s lead(s) actualizados.\n"
            "👤 %(new_contacts)s contacto(s) creados.\n"
            "� %(updated_contacts)s contacto(s) actualizados.",
            created=created,
            updated_leads=updated_leads,
            new_contacts=created_contacts,
            updated_contacts=updated_contacts,
        )

        # Agregar información sobre enriquecimiento
        if enriched_count > 0:
            summary += (
                f"\n🔍 {enriched_count} lead(s) enriquecidos con datos empresariales."
            )

        # Agregar advertencias y errores de validación
        if skipped:
            summary += "\n\n⚠️ ERRORES DE VALIDACIÓN:\n" + "\n".join(skipped[:20])
            if len(skipped) > 20:
                summary += "\n" + _(
                    "... y %(count)s errores adicionales.", count=len(skipped) - 20
                )

        # 🎯 RECALCULAR PUNTAJES Y CAPTURAR DETALLES
        scoring_details = ""
        if lead_ids_created:
            created_leads = self.env["crm.lead"].browse(lead_ids_created)
            # Forzar la recomputación del campo lead_score
            created_leads._compute_lead_score()

            # 📧 VERIFICAR Y ENVIAR ALERTAS PARA LEADS CON SCORE ALTO (>= 70)
            high_score_leads = created_leads.filtered(lambda l: l.lead_score >= 70)
            if high_score_leads:
                _logger.info(
                    f"🔔 Se encontraron {len(high_score_leads)} leads con score alto (>=70). Enviando alertas..."
                )
                for lead in high_score_leads:
                    try:
                        lead._send_high_score_alert()
                    except Exception as e:
                        _logger.error(
                            f"Error enviando alerta para lead {lead.name}: {str(e)}"
                        )

            # Calcular estadísticas
            high_score = len([l for l in created_leads if l.lead_score >= 70])
            medium_score = len([l for l in created_leads if 50 <= l.lead_score < 70])
            low_score = len([l for l in created_leads if l.lead_score < 50])
            avg_score = (
                sum([l.lead_score for l in created_leads]) / len(created_leads)
                if created_leads
                else 0
            )

            # Crear tabla HTML minimalista y profesional
            scoring_details = f"""
<!-- Métricas KPI - Estilo minimalista -->
<div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; margin-bottom: 24px;">
    <div style="background: #f8f9fa; border-left: 4px solid #6c63ff; padding: 16px; border-radius: 4px;">
        <div style="font-size: 11px; font-weight: 600; color: #6c757d; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Score Promedio</div>
        <div style="font-size: 28px; font-weight: 700; color: #212529;">{avg_score:.1f}</div>
    </div>
    <div style="background: #f8f9fa; border-left: 4px solid #10b981; padding: 16px; border-radius: 4px;">
        <div style="font-size: 11px; font-weight: 600; color: #6c757d; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Alto (≥70)</div>
        <div style="font-size: 28px; font-weight: 700; color: #10b981;">{high_score}</div>
    </div>
    <div style="background: #f8f9fa; border-left: 4px solid #f59e0b; padding: 16px; border-radius: 4px;">
        <div style="font-size: 11px; font-weight: 600; color: #6c757d; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Medio (50-69)</div>
        <div style="font-size: 28px; font-weight: 700; color: #f59e0b;">{medium_score}</div>
    </div>
    <div style="background: #f8f9fa; border-left: 4px solid #ef4444; padding: 16px; border-radius: 4px;">
        <div style="font-size: 11px; font-weight: 600; color: #6c757d; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Bajo (&lt;50)</div>
        <div style="font-size: 28px; font-weight: 700; color: #ef4444;">{low_score}</div>
    </div>
</div>

<!-- Tabla de datos - Estilo corporativo limpio -->
<table style="width: 100%; border-collapse: separate; border-spacing: 0; background: white; border: 1px solid #e5e7eb; border-radius: 6px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.05);">
    <thead>
        <tr style="background: #f9fafb; border-bottom: 2px solid #e5e7eb;">
            <th style="padding: 14px 16px; text-align: left; font-size: 12px; font-weight: 600; color: #374151; text-transform: uppercase; letter-spacing: 0.5px;">Lead</th>
            <th style="padding: 14px 16px; text-align: left; font-size: 12px; font-weight: 600; color: #374151; text-transform: uppercase; letter-spacing: 0.5px;">Contacto</th>
            <th style="padding: 14px 16px; text-align: center; font-size: 12px; font-weight: 600; color: #374151; text-transform: uppercase; letter-spacing: 0.5px; width: 90px;">Score</th>
            <th style="padding: 14px 16px; text-align: left; font-size: 12px; font-weight: 600; color: #374151; text-transform: uppercase; letter-spacing: 0.5px;">Monto</th>
            <th style="padding: 14px 16px; text-align: left; font-size: 12px; font-weight: 600; color: #374151; text-transform: uppercase; letter-spacing: 0.5px;">País</th>
            <th style="padding: 14px 16px; text-align: left; font-size: 12px; font-weight: 600; color: #374151; text-transform: uppercase; letter-spacing: 0.5px;">Fuente</th>
            <th style="padding: 14px 16px; text-align: left; font-size: 12px; font-weight: 600; color: #374151; text-transform: uppercase; letter-spacing: 0.5px;">Criterios</th>
        </tr>
    </thead>
    <tbody>
"""

            for idx, lead in enumerate(created_leads):
                # Determinar color del score
                if lead.lead_score >= 70:
                    score_color = "#10b981"
                    score_bg = "#d1fae5"
                elif lead.lead_score >= 50:
                    score_color = "#f59e0b"
                    score_bg = "#fef3c7"
                elif lead.lead_score >= 30:
                    score_color = "#f97316"
                    score_bg = "#ffedd5"
                else:
                    score_color = "#ef4444"
                    score_bg = "#fee2e2"

                # Formatear monto
                monto_str = (
                    f"${lead.expected_revenue:,.0f}"
                    if lead.expected_revenue
                    else "Sin monto"
                )

                # País
                pais_str = lead.country_id.name if lead.country_id else "—"

                # Fuente
                fuente_str = lead.source_id.name if lead.source_id else "—"

                # Obtener criterios de scoring
                try:
                    score_details = lead.get_scoring_details()
                    if score_details:
                        # Crear badges limpios para cada criterio
                        criteria_badges = []
                        for detail in score_details:
                            # Determinar color del badge según puntos
                            if detail["points"] >= 50:
                                badge_color = "#10b981"
                            elif detail["points"] >= 20:
                                badge_color = "#6366f1"
                            elif detail["points"] >= 10:
                                badge_color = "#8b5cf6"
                            else:
                                badge_color = "#6b7280"

                            criteria_badges.append(
                                f'<span style="display: inline-block; margin: 2px 4px 2px 0; padding: 3px 8px; '
                                f"background-color: {badge_color}; color: white; border-radius: 3px; "
                                f'font-size: 10px; font-weight: 500; white-space: nowrap;">'
                                f'{detail["rule"].replace("✅ ", "").replace("🔥 ", "").replace("⭐ ", "").replace("⚡ ", "").replace("🇺🇸 ", "").replace("🇲🇽 ", "").replace("🇪🇸 ", "").replace("🏢 ", "").replace("👥 ", "").replace("📧 ", "").replace("🌐 ", "").replace("📱 ", "")} '
                                f'<strong>+{detail["points"]}</strong></span>'
                            )
                        criteria_html = "".join(criteria_badges)
                    else:
                        criteria_html = '<span style="color: #9ca3af; font-size: 12px;">Sin criterios</span>'
                except Exception as e:
                    _logger.error(
                        f"Error getting scoring details for lead {lead.id}: {str(e)}"
                    )
                    criteria_html = (
                        '<span style="color: #ef4444; font-size: 12px;">Error</span>'
                    )

                # Alternar color de fondo para filas (zebra striping sutil)
                row_bg = "#ffffff" if idx % 2 == 0 else "#f9fafb"

                # Agregar fila a la tabla
                scoring_details += f"""
        <tr style="background: {row_bg}; border-bottom: 1px solid #e5e7eb;">
            <td style="padding: 12px 16px;">
                <div style="font-weight: 600; color: #111827; font-size: 13px;">{lead.name}</div>
            </td>
            <td style="padding: 12px 16px;">
                <div style="font-size: 12px; color: #6b7280;">{lead.email_from or '—'}</div>
                <div style="font-size: 12px; color: #9ca3af; margin-top: 2px;">{lead.phone or '—'}</div>
            </td>
            <td style="padding: 12px 16px; text-align: center;">
                <div style="display: inline-block; padding: 6px 12px; background: {score_bg}; border-radius: 4px;">
                    <span style="font-size: 18px; font-weight: 700; color: {score_color};">{lead.lead_score}</span>
                </div>
            </td>
            <td style="padding: 12px 16px;">
                <div style="font-weight: 600; color: #111827; font-size: 13px;">{monto_str}</div>
            </td>
            <td style="padding: 12px 16px;">
                <div style="font-size: 12px; color: #6b7280;">{pais_str}</div>
            </td>
            <td style="padding: 12px 16px;">
                <div style="font-size: 12px; color: #6b7280;">{fuente_str}</div>
            </td>
            <td style="padding: 12px 16px;">
                <div style="display: flex; flex-wrap: wrap; gap: 2px;">
                    {criteria_html}
                </div>
            </td>
        </tr>
"""

            scoring_details += """
    </tbody>
</table>
"""

        # 🔍 GENERAR TABLA HTML DE ENRIQUECIMIENTO
        enrichment_details = ""
        if enrichment_log:
            enrichment_details = f"""
<!-- Encabezado de Enriquecimiento -->
<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 24px; margin-bottom: 24px; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
    <h3 style="color: #ffffff; margin: 0; font-weight: 700; font-size: 20px; letter-spacing: -0.5px;">
        🔍 Información Enriquecida de Empresas
    </h3>
    <p style="color: rgba(255,255,255,0.9); margin: 8px 0 0 0; font-size: 14px;">
        {len(enrichment_log)} lead(s) con información empresarial automática
    </p>
</div>
"""

            for idx, enriched_item in enumerate(enrichment_log):
                data = enriched_item["enriched_data"]

                # Extraer información del lead
                nombre = enriched_item["nombre"]
                email = enriched_item["email"]
                telefono = enriched_item.get("telefono", "No proporcionado")
                domain = email.split("@")[1] if "@" in email else ""

                # Extraer información enriquecida de la empresa
                company_name = data.get("company_name", domain.split(".")[0].title())
                country = data.get("country", "No disponible")
                industry = data.get("industry", "No especificada")
                employees = data.get("employee_count", "No disponible")
                description = data.get(
                    "description",
                    f"Empresa del sector {industry} ubicada en {country}.",
                )

                # Formatear número de empleados
                if employees != "No disponible" and employees:
                    if isinstance(employees, int):
                        employees = f"{employees:,}"

                # Tarjeta de información del lead y empresa
                card_bg = "#f8f9fa" if idx % 2 == 0 else "#ffffff"

                enrichment_details += f"""
<!-- Tarjeta de Lead #{idx + 1} -->
<div style="background: {card_bg}; border: 1px solid #e5e7eb; border-radius: 8px; padding: 20px; margin-bottom: 16px; box-shadow: 0 2px 4px rgba(0,0,0,0.05);">
    
    <!-- Sección: Información del Contacto -->
    <div style="border-bottom: 2px solid #e5e7eb; padding-bottom: 12px; margin-bottom: 16px;">
        <h4 style="color: #1f2937; margin: 0 0 12px 0; font-size: 16px; font-weight: 600; display: flex; align-items: center;">
            👤 Información del Contacto
        </h4>
        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 12px;">
            <div>
                <div style="font-size: 11px; font-weight: 600; color: #6b7280; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Nombre</div>
                <div style="font-size: 14px; font-weight: 600; color: #111827;">{nombre}</div>
            </div>
            <div>
                <div style="font-size: 11px; font-weight: 600; color: #6b7280; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Correo Electrónico</div>
                <div style="font-size: 13px; color: #4f46e5; font-family: monospace;">📧 {email}</div>
            </div>
            <div>
                <div style="font-size: 11px; font-weight: 600; color: #6b7280; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Teléfono</div>
                <div style="font-size: 13px; color: #059669; font-weight: 500;">📞 {telefono}</div>
            </div>
        </div>
    </div>
    
    <!-- Sección: Información de la Empresa -->
    <div>
        <h4 style="color: #1f2937; margin: 0 0 12px 0; font-size: 16px; font-weight: 600; display: flex; align-items: center;">
            🏢 Información de la Empresa
        </h4>
        
        <!-- Nombre de la empresa destacado -->
        <div style="background: #ffffff; border-left: 4px solid #6366f1; padding: 12px 16px; margin-bottom: 16px; border-radius: 4px;">
            <div style="font-size: 11px; font-weight: 600; color: #6b7280; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Empresa</div>
            <div style="font-size: 18px; font-weight: 700; color: #111827;">{company_name}</div>
            <div style="font-size: 11px; color: #9ca3af; margin-top: 2px; font-family: monospace;">🌐 {domain}</div>
        </div>
        
        <!-- Grid con datos de la empresa -->
        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 12px; margin-bottom: 16px;">
            <div style="background: #f0fdf4; padding: 12px; border-radius: 6px; border: 1px solid #bbf7d0;">
                <div style="font-size: 11px; font-weight: 600; color: #065f46; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">País</div>
                <div style="font-size: 14px; font-weight: 600; color: #047857;">🌍 {country}</div>
            </div>
            <div style="background: #fef3c7; padding: 12px; border-radius: 6px; border: 1px solid #fde68a;">
                <div style="font-size: 11px; font-weight: 600; color: #92400e; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Industria</div>
                <div style="font-size: 14px; font-weight: 600; color: #b45309;">🏭 {industry}</div>
            </div>
            <div style="background: #ede9fe; padding: 12px; border-radius: 6px; border: 1px solid #ddd6fe;">
                <div style="font-size: 11px; font-weight: 600; color: #5b21b6; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Empleados</div>
                <div style="font-size: 14px; font-weight: 600; color: #6d28d9;">👥 {employees}</div>
            </div>
        </div>
        
        <!-- Descripción de la empresa -->
        <div style="background: #ffffff; padding: 16px; border-radius: 6px; border: 1px solid #e5e7eb;">
            <div style="font-size: 11px; font-weight: 600; color: #6b7280; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 8px;">📝 Descripción</div>
            <div style="font-size: 13px; color: #374151; line-height: 1.6;">{description}</div>
        </div>
    </div>
</div>
"""

            enrichment_details += """
<div style="margin-top: 24px; padding: 16px; background: #f9fafb; border-radius: 6px; border: 1px solid #e5e7eb; text-align: center;">
    <div style="font-size: 12px; color: #6b7280;">
        💡 <strong>Datos enriquecidos automáticamente</strong> mediante APIs externas (Clearbit, Hunter.io)
    </div>
</div>
"""

        # Actualizar los campos del registro
        self.write(
            {
                "result_message": summary,
                "scoring_details": scoring_details,
                "enrichment_details": enrichment_details,
                "state": "done",
                "leads_count": created,
                "leads_updated": updated_leads,
                "contacts_created": created_contacts,
                "contacts_updated": updated_contacts,
                "lead_ids": [(6, 0, lead_ids_created)],
            }
        )

        # Actualizar el nombre si sigue siendo el predeterminado
        if "Importación #" in self.name:
            self.name = f'{self.name} - {self.file_name or "Sin nombre"}'

        # Registrar en el chatter: usar message_notify para notificar al usuario actual (evita error en modelos transient)
        try:
            partner_id = (
                self.env.user.partner_id.id
                if self.env.user and self.env.user.partner_id
                else False
            )
            if partner_id:
                self.message_notify(
                    partner_ids=[partner_id],
                    body=_(
                        "Importación completada: %(leads)s leads creados.",
                        leads=created,
                    ),
                    subject="Importación de Leads",
                )
            else:
                # Fallback: registrar como mensaje en el registro si es seguro
                self.message_post(
                    body=_(
                        "Importación completada: %(leads)s leads creados.",
                        leads=created,
                    )
                )
        except Exception:
            _logger.exception("No se pudo notificar la finalización de la importación")

        # Mostrar mensaje de éxito
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("✅ Importación Completada"),
                "message": _(
                    "Se importaron %(count)s leads correctamente.", count=created
                ),
                "type": "success",
                "sticky": False,
                "next": {"type": "ir.actions.act_window_close"},
            },
        }

    def _read_rows(self, file_content, extension):
        # Lee las filas del archivo según la extensión
        if extension == ".csv":
            return self._read_csv(file_content)
        if extension == ".xlsx":
            return self._read_xlsx(file_content)
        if extension == ".xls":
            return self._read_xls(file_content)
        return []

    def _read_csv(self, file_content):
        # Lee y valida un archivo CSV, devolviendo las filas como diccionarios
        text = file_content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        self._validate_headers(headers)
        rows = []
        for row in reader:
            if not row or not any(
                value.strip() for value in row.values() if isinstance(value, str)
            ):
                continue
            rows.append(
                {key.strip(): (value or "").strip() for key, value in row.items()}
            )
        return rows

    def _read_xlsx(self, file_content):
        # Lee y valida un archivo XLSX, devolviendo las filas como diccionarios
        if not openpyxl:
            raise UserError(_("El servidor no tiene el paquete openpyxl instalado."))
        workbook = openpyxl.load_workbook(
            io.BytesIO(file_content), read_only=True, data_only=True
        )
        sheet = workbook.active
        headers = []
        rows = []
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx == 0:
                headers = [self._to_header(cell) for cell in row]
                self._validate_headers(headers)
                continue
            if not row or not any(row):
                continue
            row_data = {}
            for col, header in enumerate(headers):
                if not header:
                    continue
                value = row[col] if col < len(row) else ""
                row_data[header] = self._to_cell_value(value)
            rows.append(row_data)
        return rows

    def _read_xls(self, file_content):
        # Lee y valida un archivo XLS, devolviendo las filas como diccionarios
        if not xlrd:
            raise UserError(_("El servidor no tiene el paquete xlrd instalado."))
        workbook = xlrd.open_workbook(file_contents=file_content)
        sheet = workbook.sheet_by_index(0)
        headers = [
            self._to_header(sheet.cell_value(0, col)) for col in range(sheet.ncols)
        ]
        self._validate_headers(headers)
        rows = []
        for row_idx in range(1, sheet.nrows):
            values = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(
                        value, workbook.datemode
                    ).date()
                values.append(value)
            if not values or not any(values):
                continue
            row_data = {}
            for col, header in enumerate(headers):
                if not header:
                    continue
                value = values[col] if col < len(values) else ""
                row_data[header] = self._to_cell_value(value)
            rows.append(row_data)
        return rows

    def _validate_headers(self, headers):
        # Valida que el archivo tenga todos los encabezados obligatorios
        normalized = [header.strip() for header in headers if header]
        missing = [
            header for header in self._required_headers if header not in normalized
        ]
        if missing:
            raise UserError(
                _(
                    "Faltan columnas obligatorias: %(columns)s",
                    columns=", ".join(missing),
                )
            )

    def _to_header(self, cell):
        # Normaliza el valor de la celda para usarlo como encabezado
        if cell is None:
            return ""
        return str(cell).strip()

    def _to_cell_value(self, cell):
        # Normaliza el valor de la celda para usarlo como dato
        if cell is None:
            return ""
        if isinstance(cell, (bytes, bytearray)):
            return cell.decode("utf-8", errors="ignore")
        return str(cell).strip()

    def _get_or_create_partner_sql(self, row):
        # Busca o crea el contacto (partner) usando SQL directo
        name = row.get("Nombre") or ""
        email = row.get("Correo") or ""
        phone = row.get("Telefono") or ""
        phone_2 = row.get("Teléfono 2") or ""
        country_name = row.get("Pais") or ""
        city = row.get("Ciudad") or ""

        if not name:
            raise UserError(_("El contacto debe tener un nombre."))

        try:
            country_id = None
            if country_name:
                # name es JSONB, se castea a texto para comparación ILIKE
                self.env.cr.execute(
                    "SELECT id FROM res_country WHERE (name->>'es_MX')::text ILIKE %s OR (name->>'en_US')::text ILIKE %s OR code ILIKE %s LIMIT 1",
                    (
                        f"%{country_name.strip()}%",
                        f"%{country_name.strip()}%",
                        f"{country_name.strip()}%",
                    ),
                )
                res = self.env.cr.fetchone()
                if res:
                    country_id = res[0]

            partner_id = None
            if email:
                self.env.cr.execute(
                    "SELECT id FROM res_partner WHERE email = %s LIMIT 1", (email,)
                )
                res = self.env.cr.fetchone()
                if res:
                    partner_id = res[0]

            if not partner_id and phone:
                self.env.cr.execute(
                    "SELECT id FROM res_partner WHERE phone = %s OR mobile = %s LIMIT 1",
                    (phone, phone),
                )
                res = self.env.cr.fetchone()
                if res:
                    partner_id = res[0]

            if not partner_id and phone_2:
                self.env.cr.execute(
                    "SELECT id FROM res_partner WHERE phone = %s OR mobile = %s LIMIT 1",
                    (phone_2, phone_2),
                )
                res = self.env.cr.fetchone()
                if res:
                    partner_id = res[0]

            if partner_id:
                _logger.info(
                    "Partner found by SQL with id %s for email/phone %s/%s",
                    partner_id,
                    email,
                    phone,
                )
                return partner_id, {"created": 0, "updated": 1}

            # Crear el contacto si no existe
            self.env.cr.execute(
                """INSERT INTO res_partner 
                (name, email, phone, mobile, city, country_id, active, autopost_bills, create_uid, create_date, write_uid, write_date)
                VALUES (%s, %s, %s, %s, %s, %s, true, false, %s, now(), %s, now())
                RETURNING id""",
                (
                    name or "",
                    email or None,
                    phone or None,
                    phone_2 or phone or None,
                    city or None,
                    country_id,
                    self.env.uid,
                    self.env.uid,
                ),
            )
            res = self.env.cr.fetchone()
            new_id = res[0] if res else None

            if new_id:
                _logger.info(
                    "Partner created by SQL with id %s for name/email %s/%s",
                    new_id,
                    name,
                    email,
                )
                return new_id, {"created": 1, "updated": 0}
            else:
                raise ValueError("Failed to create partner via SQL")

        except Exception as err:
            _logger.exception("Partner creation failed via SQL: %s", err)
            raise UserError(
                _("Error al crear/buscar contacto: %(error)s", error=str(err))
            )

    def _create_lead_sql(self, row, partner_id):
        # Crea el lead usando ORM, asociando el contacto y la fuente
        name = row.get("Nombre") or ""
        email = row.get("Correo") or ""
        phone = row.get("Telefono") or ""
        phone_2 = row.get("Teléfono 2") or ""
        company = row.get("Empresa") or ""
        country_name = row.get("Pais") or ""
        city = row.get("Ciudad") or ""
        source_name = row.get("Fuente") or ""
        medium_name = row.get("Marca campaña") or ""
        campaign_name = row.get("Nombre campaña") or ""
        program_interest = row.get("Curso / Programa interés") or ""
        profile = row.get("Perfil") or ""
        observations = row.get("Observaciones") or ""
        monto = row.get("Monto") or "0"

        try:
            expected_revenue = 0.0
            if monto:
                try:
                    expected_revenue = float(monto)
                except ValueError:
                    expected_revenue = 0.0

            country_id = self._find_country(country_name)

            stage_new = self.env["crm.stage"].search(
                [
                    ("name", "=", "Nuevo"),
                    ("team_id", "=", False),
                ],
                order="sequence",
                limit=1,
            )

            lead_vals = {
                "name": company or name,
                "contact_name": name,
                "type": "opportunity",
                "partner_id": partner_id or None,
                "email_from": email or None,
                "phone": phone or None,
                "mobile": phone_2 or None,
                "expected_revenue": expected_revenue,
                "stage_id": stage_new.id if stage_new else False,
                "program_interest": program_interest or False,
                "profile": self._map_profile(profile) or False,
                "description": observations or False,
            }

            if country_id:
                lead_vals["country_id"] = country_id
            if city:
                lead_vals["city"] = city
            if source_name:
                lead_vals["source_id"] = self._get_or_create_utm_source(source_name)
            if medium_name:
                lead_vals["medium_id"] = self._get_or_create_utm_medium(medium_name)
            if campaign_name:
                lead_vals["campaign_id"] = self._get_or_create_utm_campaign(
                    campaign_name
                )

            lead = self.env["crm.lead"].create(lead_vals)
            _logger.info("Lead creado por ORM con id %s para %s", lead.id, lead.name)
            return lead.id

        except Exception as err:
            _logger.exception("Lead creation failed via ORM: %s", err)
            raise UserError(_("Error al crear lead: %(error)s", error=str(err)))

    def _find_country(self, country_name):
        # Busca el país por nombre o código usando SQL directo
        if not country_name:
            return False
        try:
            # name es JSONB, se castea a texto para comparación ILIKE
            self.env.cr.execute(
                "SELECT id FROM res_country WHERE (name->>'es_MX')::text ILIKE %s OR (name->>'en_US')::text ILIKE %s OR code ILIKE %s LIMIT 1",
                (
                    f"%{country_name.strip()}%",
                    f"%{country_name.strip()}%",
                    f"{country_name.strip()}%",
                ),
            )
            res = self.env.cr.fetchone()
            if res:
                return res[0]
            return False
        except Exception:
            _logger.exception("Error finding country: %s", country_name)
            return False

    def _get_or_create_utm_source(self, source_name):
        """Obtiene o crea una fuente UTM por nombre."""
        name = (source_name or "").strip()
        if not name:
            return False
        source = self.env["utm.source"].search([("name", "=", name)], limit=1)
        if not source:
            source = self.env["utm.source"].create({"name": name})
        return source.id

    def _get_or_create_utm_medium(self, medium_name):
        """Obtiene o crea un medio UTM por nombre."""
        name = (medium_name or "").strip()
        if not name:
            return False
        medium = self.env["utm.medium"].search([("name", "=", name)], limit=1)
        if not medium:
            medium = self.env["utm.medium"].create({"name": name})
        return medium.id

    def _get_or_create_utm_campaign(self, campaign_name):
        """Obtiene o crea una campaña UTM por nombre."""
        name = (campaign_name or "").strip()
        if not name:
            return False
        campaign = self.env["utm.campaign"].search([("name", "=", name)], limit=1)
        if not campaign:
            campaign = self.env["utm.campaign"].create({"name": name})
        return campaign.id

    def _map_profile(self, profile_raw):
        """
        Normaliza el valor de 'Perfil' del CSV a las claves aceptadas por el campo
        selection `profile` en `crm.lead`.

        Devuelve la clave de selección (ej: 'estudiante') o False si no se puede
        mapear.
        """
        if not profile_raw:
            return False
        # Normalizar: minusculas, sin espacios externos, sin acentos
        import unicodedata

        def _strip_accents(s):
            nk = unicodedata.normalize("NFKD", s)
            return "".join(c for c in nk if not unicodedata.combining(c))

        val = _strip_accents(str(profile_raw).strip().lower())

        mapping = {
            "estudiante": "estudiante",
            "estudiantes": "estudiante",
            "profesional": "profesional",
            "profesionales": "profesional",
            "empresario": "empresario",
            "empresaria": "empresario",
            "empleado": "empleado",
            "empleada": "empleado",
            "independiente": "independiente",
            "independentes": "independiente",
            # variantes comunes / cargos -> mapear a la categoría más apropiada
            "ejecutivo": "empleado",
            "perfil ejecutivo": "empleado",
            "ejecutiva": "empleado",
            "emprendedor": "empresario",
            "emprendedora": "empresario",
            "ingeniero": "profesional",
            "ingeniera": "profesional",
            "vendedor": "profesional",
            "vendedora": "profesional",
            "disenadora": "profesional",
            "disenador": "profesional",
            "rrhh": "empleado",
            # fallback words
            "otro": "otro",
        }

        # If exact mapping exists
        if val in mapping:
            return mapping[val]

        # Try to match by keywords inside the value
        if "estudiant" in val:
            return "estudiante"
        if "profes" in val or "ingenier" in val or "disen" in val or "vended" in val:
            return "profesional"
        if "empres" in val or "emprend" in val:
            return "empresario"
        if "independ" in val:
            return "independiente"
        if "ejecut" in val or "rrhh" in val or "emple" in val:
            return "empleado"

        return False

    def _prepare_lead_vals(self, row):
        # Prepara los valores para crear un lead y maneja la creación/actualización del contacto
        # Similar a la lógica de importación pero para uso programático
        name = row.get("Nombre") or ""
        email = row.get("Correo") or ""
        phone = row.get("Telefono") or ""
        phone_2 = row.get("Teléfono 2") or ""
        company = row.get("Empresa") or ""
        country_name = row.get("Pais") or ""
        city = row.get("Ciudad") or ""
        source_name = row.get("Fuente") or ""
        medium_name = row.get("Marca campaña") or ""
        campaign_name = row.get("Nombre campaña") or ""
        program_interest = row.get("Curso / Programa interés") or ""
        profile = row.get("Perfil") or ""
        observations = row.get("Observaciones") or ""

        # Buscar o crear partner
        partner_id, partner_stats = self._get_or_create_partner_sql(row)

        # Preparar valores para el lead
        lead_vals = {
            "name": company or name,
            "contact_name": name,
            "type": "lead",
            "partner_id": partner_id,
            "email_from": email,
            "phone": phone,
            "mobile": phone_2 or False,
            "program_interest": program_interest or False,
            "profile": self._map_profile(profile) or False,
            "description": observations or False,
        }

        # País
        if country_name:
            country_id = self._find_country(country_name)
            if country_id:
                lead_vals["country_id"] = country_id
        if city:
            lead_vals["city"] = city

        # Fuente UTM
        if source_name:
            lead_vals["source_id"] = self._get_or_create_utm_source(source_name)
        if medium_name:
            lead_vals["medium_id"] = self._get_or_create_utm_medium(medium_name)
        if campaign_name:
            lead_vals["campaign_id"] = self._get_or_create_utm_campaign(campaign_name)

        return lead_vals, partner_stats

    def _detect_extension(self, file_content):
        # Detecta la extensión del archivo por su cabecera binaria
        if not file_content:
            return ""
        if file_content[:2] == b"PK":
            return ".xlsx"
        if file_content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            return ".xls"
        return ".csv"
